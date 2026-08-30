import streamlit as st
import streamlit.components.v1 as components
import sqlite3, time, calendar as cal_lib, glob, os, re, threading, json, io, hashlib
import pandas as pd
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

TT_TOOL = os.path.join(os.path.dirname(__file__), "시간표도구", "개강안내.html")

def _tt_tool_html():
    # 파일 mtime 을 캐시 키에 넣어, 재배포로 개강안내.html 이 바뀌면 콜드 재시작 없이도 새로 읽는다
    try:
        _m = os.path.getmtime(TT_TOOL)
    except OSError:
        _m = 0
    return _tt_tool_html_cached(_m)

@st.cache_data
def _tt_tool_html_cached(_mtime):
    with open(TT_TOOL, encoding="utf-8") as f:
        return f.read()

def _consultant_inject(consultant):
    """window.__CONSULTANT__ 주입 스크립트 조각 (개강안내.html consultantInfo() 가 읽음)."""
    if not consultant:
        return ""
    js = json.dumps(consultant, ensure_ascii=False).replace("</", "<\\/")
    return "try{window.__CONSULTANT__=" + js + ";}catch(e){}\n"

def render_timetable_tool(hash_tab="timetable", preload=None, height=1500, consultant=None):
    """클로드놀이_배포판 개강안내.html 을 그대로 embed.
    hash_tab: 'timetable' | 'assign' | 'announce' | 'consult' | 'upload'
    preload: [(name, content_b64), ...] — 도구에 미리 주입(재업로드 불필요).
    consultant: {name,phone,mobile,email,date} — 문서 서명(담당자) 값. 없으면 도구 하드코딩 fallback.
    ※ 4개 메뉴가 각각 별도 iframe이라 상태가 안 섞임 → preload(DB 저장분)로 모두 같은 시간표를 받는다."""
    html = _tt_tool_html()
    inject = "<script>(function(){\n" + _consultant_inject(consultant)
    if preload:
        files_js = json.dumps([{"name": n, "b64": b} for n, b in preload], ensure_ascii=False)
        inject += (
            "var FS=" + files_js + ";\n"
            "async function _pre(){try{\n"
            "  if(typeof loadFileAuto!=='function'){setTimeout(_pre,150);return;}\n"
            "  for(var i=0;i<FS.length;i++){var b=atob(FS[i].b64),a=new Uint8Array(b.length);\n"
            "    for(var k=0;k<b.length;k++)a[k]=b.charCodeAt(k);\n"
            "    await loadFileAuto(new File([a],FS[i].name));}\n"
            "  if(typeof saveToStorage==='function')saveToStorage();\n"
            "  if(typeof updateMonthFilters==='function')updateMonthFilters();\n"
            "  if(typeof updateAllMonthFilters==='function')updateAllMonthFilters();\n"
            "  location.hash='#" + hash_tab + "';\n"
            "  if(typeof applyHashTab==='function')applyHashTab();\n"     # 데이터 채운 뒤 해당 탭 다시 렌더
            "  else if(typeof switchTab==='function')switchTab('" + hash_tab + "');\n"
            "}catch(e){console.warn('시간표 preload 실패',e);}}\n"
            "if(document.readyState==='complete')_pre();else window.addEventListener('load',function(){setTimeout(_pre,200);});\n")
    else:
        inject += (
            "function _go(){location.hash='#" + hash_tab + "';\n"
            "  if(typeof applyHashTab==='function')applyHashTab();}\n"
            "if(document.readyState==='complete')_go();else window.addEventListener('load',_go);\n")
    inject += "})();</script>"
    # ※ '</body>'·'</html>'는 도구 JS의 인쇄/PNG 템플릿 문자열 안에도 있어 replace 불가 → 문서 맨 끝에 덧붙임.
    components.html(html + inject, height=height, scrolling=True)

DB = "salesdb.db"
TT_DIR = "인트라넷 시간표"
FEE_FILE = "수강료.xlsx"
DAYS = "월화수목금토일"
KST = ZoneInfo("Asia/Seoul")

# ── 시간표 엑셀(HTML export) 파싱 ────────────────────────────────
CELL_RE = re.compile(
    r'^(?P<subject>.+?)\s*전체출석율\s*:\s*[\d.]+%.*?'
    r'정원\s*:\s*(?P<cap>\d+)\((?P<enrolled>\d+)\).*?'
    r'배정\s*:\s*(?P<assigned>\d+)(?:\(W:\d+,R:\d+\))?'
    r'(?P<rest>.+?)'
    r'개\s*:\s*(?P<start_date>\d{4}-\d{2}-\d{2})종\s*:\s*(?P<end_date>\d{4}-\d{2}-\d{2})\s*$'
)
DAY_TAIL_RE = re.compile(r'([월화수목금토일][월화수목금토일~,/]*)$')

def _end_plus30(t):
    h, m = map(int, t.split(":"))
    m += 30
    if m >= 60: m -= 60; h += 1
    return f"{h:02d}:{m:02d}"

def _expand_days(tok):
    tok = tok.strip()
    if "~" in tok:
        a, b = tok.split("~")
        if a in DAYS and b in DAYS:
            i, j = DAYS.index(a), DAYS.index(b)
            return list(DAYS[i:j+1])
        return []
    seen = []
    for ch in tok:
        if ch in DAYS and ch not in seen:
            seen.append(ch)
    return seen

def _is_weekend_days(days):
    return bool(days) and set(days) <= {"토", "일"}

def _parse_cell(text, room, time_label):
    m = CELL_RE.match(text.strip())
    if not m:
        return None
    rest = m.group("rest")
    dm = DAY_TAIL_RE.search(rest)
    if not dm:
        return None
    day_tok = dm.group(1)
    days = _expand_days(day_tok)
    if not days:
        return None
    teacher = re.sub(r"\d+$", "", rest[:dm.start()].strip())
    teacher = re.sub(r"^재직자\s*:\s*\d+", "", teacher).strip()
    subject = re.sub(r"/(주말|격일)$", "", m.group("subject")).strip()
    return dict(room=room, subject=subject, teacher=teacher, days=days, day_label=day_tok,
                start_date=m.group("start_date"), end_date=m.group("end_date"),
                cap=int(m.group("cap")), enrolled=int(m.group("enrolled")),
                assigned=int(m.group("assigned")), start_time=time_label)

@st.cache_data
def _tt_folder_b64():
    """`인트라넷 시간표/` 폴더의 xls 를 (name, base64) 로. 임베드 도구 preload 폴백용 (저장소에 커밋돼 팀원 모두 동일)."""
    import base64
    out = []
    for p in sorted(glob.glob(os.path.join(TT_DIR, "*.xls"))):
        try:
            with open(p, "rb") as f:
                out.append((os.path.basename(p), base64.b64encode(f.read()).decode()))
        except Exception:
            pass
    return out

def _tt_source_htmls():
    """시간표 원본 (name, html_text) 목록. DB 업로드분이 있으면 그것만, 없으면 `인트라넷 시간표/` 폴더. (캐시 안 함 — 업로드 반영)"""
    import base64
    try:
        ups = q("SELECT name, content_b64 FROM tt_upload ORDER BY name")
    except Exception:
        ups = []
    if ups:
        out = []
        for name, b64 in ups:
            try:
                out.append((name, base64.b64decode(b64).decode("utf-8", "replace")))
            except Exception:
                pass
        return out
    out = []
    for p in sorted(glob.glob(os.path.join(TT_DIR, "*.xls"))):
        try:
            with open(p, encoding="utf-8", errors="replace") as f:
                out.append((os.path.basename(p), f.read()))
        except Exception:
            pass
    return out

@st.cache_data
def _load_timetable(_cache_key, htmls):
    sessions = []
    seen = set()  # 같은 강좌가 여러 스냅샷 파일에 중복 등록되는 걸 걸러냄
    for _name, html in htmls:
        try:
            df = pd.read_html(io.StringIO(html), header=[0, 1])[0]
        except Exception:
            continue
        df = df[df.iloc[:, 0] != "정원"].reset_index(drop=True)
        times = df.iloc[:, 0].tolist()
        for col in df.columns[1:]:
            room = col[0]
            vals = df[col].tolist()
            i = 0
            while i < len(vals):
                v = vals[i]
                if pd.isna(v):
                    i += 1; continue
                j = i
                while j + 1 < len(vals) and vals[j + 1] == v:
                    j += 1
                sess = _parse_cell(str(v), room, times[i])
                if sess:
                    sess["end_time"] = _end_plus30(times[j])
                    dedup_key = (sess["subject"], sess["room"], sess["teacher"], sess["day_label"],
                                 sess["start_date"], sess["end_date"], sess["start_time"])
                    if dedup_key not in seen:
                        seen.add(dedup_key)
                        sessions.append(sess)
                i = j + 1
    return sessions

def get_timetable():
    htmls = tuple(_tt_source_htmls())
    key = tuple((n, hashlib.md5(h.encode("utf-8", "replace")).hexdigest()) for n, h in htmls)
    return _load_timetable(key, htmls)

# ── 개인 시간표(양식 xlsx) 파싱 ───────────────────────────────────
def _parse_personal_cell(text):
    lines = [l.strip() for l in str(text).split("\n") if l.strip()]
    if not lines:
        return None
    start_date = end_date = None
    day_tok = None
    for l in lines:
        m = re.match(r"^개\s*:\s*(\d{4}-\d{2}-\d{2})$", l)
        if m: start_date = m.group(1)
        m = re.match(r"^종\s*:\s*(\d{4}-\d{2}-\d{2})$", l)
        if m: end_date = m.group(1)
        dm = re.match(r"^[월화수목금토일][월화수목금토일~,/]*$", l)
        if dm: day_tok = l
    if not (start_date and end_date and day_tok):
        return None
    days = _expand_days(day_tok)
    if not days:
        return None
    return dict(subject=lines[0], day_label=day_tok, days=days,
                start_date=start_date, end_date=end_date)

def parse_personal_timetable(file):
    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    student = ""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                m = re.search(r"(\S+)님\s*개인\s*시간표", cell.value)
                if m:
                    student = m.group(1).strip()
                    break
        if student:
            break

    courses = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows:
        cells = list(row)
        if not cells or cells[0] == "비고":
            continue
        rest = [c for c in cells[1:] if c not in (None, "")]
        if not rest:
            continue
        # 월 라벨 행(예: "7월","6월"만 있는 행)은 건너뜀
        if all(isinstance(c, str) and re.match(r"^\d{1,2}월$", c) for c in rest):
            continue
        course_cells = [c for c in cells[1:] if isinstance(c, str) and "개:" in c]
        if not course_cells:
            continue
        time_lines = []
        if isinstance(cells[0], str):
            time_lines = [t.strip() for t in cells[0].split("\n") if t.strip()]
        for i, ctext in enumerate(course_cells):
            parsed = _parse_personal_cell(ctext)
            if not parsed:
                continue
            if len(time_lines) == len(course_cells):
                tl = time_lines[i]
            elif time_lines:
                tl = time_lines[0]
            else:
                tl = ""
            start_time = tl.split("~")[0].strip() if tl else ""
            parsed["start_time"] = start_time
            courses.append(parsed)
    return student, courses

# ── 시간표 이미지 인식 (Gemini vision) ────────────────────────────
@st.cache_resource
def get_gemini_client():
    api_key = st.secrets.get("GEMINI_API_KEY", "") if hasattr(st, "secrets") else ""
    if not api_key:
        return None
    from google import genai
    return genai.Client(api_key=api_key)

def parse_timetable_image(image_bytes, mime_type):
    client = get_gemini_client()
    if not client:
        raise RuntimeError("GEMINI_API_KEY가 설정되어 있지 않아요 (Streamlit Secrets 확인).")
    from google.genai import types
    prompt = (
        "이 이미지는 학원 개인 시간표야. 표 안의 각 강좌 블록에서 정보를 추출해서 "
        "JSON 배열로만 답해. 설명 문장은 절대 붙이지 마.\n"
        '형식: [{"subject":"과목명","day_label":"월~금","start_date":"YYYY-MM-DD",'
        '"end_date":"YYYY-MM-DD","start_time":"HH:MM"}]\n'
        "값을 모르면 빈 문자열로 둬."
    )
    from google.genai import errors as genai_errors
    parts = [types.Part.from_bytes(data=image_bytes, mime_type=mime_type), prompt]
    for attempt in range(3):
        try:
            resp = client.models.generate_content(model="gemini-2.5-flash-lite", contents=parts)
            break
        except genai_errors.ServerError:
            if attempt == 2:
                raise
            time.sleep(2 * (attempt + 1))  # ponytail: fixed backoff, exponential if 503s get frequent
    m = re.search(r"\[.*\]", resp.text.strip(), re.S)
    return json.loads(m.group(0)) if m else []

def available_months():
    sessions = get_timetable()
    if not sessions:
        return []
    lo = min(s["start_date"][:7] for s in sessions)
    hi = max(s["end_date"][:7] for s in sessions)
    y, m = int(lo[:4]), int(lo[5:7])
    hi_y, hi_m = int(hi[:4]), int(hi[5:7])
    months = []
    while (y, m) <= (hi_y, hi_m):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months

def sessions_in_month(month_str, weekend):
    last_day = cal_lib.monthrange(int(month_str[:4]), int(month_str[5:7]))[1]
    m_start, m_end = f"{month_str}-01", f"{month_str}-{last_day:02d}"
    out = [s for s in get_timetable()
           if s["start_date"] <= m_end and s["end_date"] >= m_start
           and _is_weekend_days(s["days"]) == weekend]
    return sorted(out, key=lambda s: (s["start_time"], s["subject"]))

def _norm_subject(s):
    return re.sub(r"\s+", "", s or "")

def match_real_session(subject, start_date):
    # 개인 시간표(xlsx/이미지)에서 뽑아낸 값은 사람이 손으로 짠 계획이라 실제
    # 인트라넷 시간표와 요일·시간 표기가 다를 수 있어서, 과목명+개강일로
    # 실제 강좌를 찾아 그 강좌의 정확한 정보(요일/시간/강의장/강사)를 대신 씀
    subject = (subject or "").strip()
    start_date = (start_date or "").strip()
    if not subject or not start_date:
        return None
    norm_subject = _norm_subject(subject)
    same_date = [s for s in get_timetable() if s["start_date"] == start_date]
    # AI 인식이 "스케치업 1"처럼 과목명과 숫자 사이에 공백을 넣는 경우가 있어서
    # 공백을 무시하고 비교함
    exact = [s for s in same_date if _norm_subject(s["subject"]) == norm_subject]
    if exact:
        return exact[0]
    partial = [s for s in same_date
               if norm_subject in _norm_subject(s["subject"]) or _norm_subject(s["subject"]) in norm_subject]
    return partial[0] if partial else None

# ── 수강료표 파싱 ────────────────────────────────────────────
def _expand_fee_name(name):
    suffix = ""
    base = name
    sm = re.search(r"(\(방학\)|/주말)$", base)
    if sm:
        suffix = sm.group(1)
        base = base[:sm.start()]
    rm = re.match(r"^(.*?)(\d+)~(\d+)$", base)
    bases = [f"{rm.group(1)}{i}" for i in range(int(rm.group(2)), int(rm.group(3)) + 1)] if rm else [base]
    out = []
    for b in bases:
        parts = [p.strip() for p in b.split(",")] if "," in b else [b]
        out.extend(p + suffix for p in parts)
    return out

@st.cache_data
def _load_fees(_cache_key):
    fees = {}
    df = pd.read_excel(FEE_FILE, header=None)
    for _, row in df.iterrows():
        for name_col, fee_col in ((3, 4), (11, 12)):
            name, fee = row.get(name_col), row.get(fee_col)
            if isinstance(name, str) and isinstance(fee, (int, float)) and not pd.isna(fee):
                for key in _expand_fee_name(name.strip()):
                    fees[key] = int(fee)
    return fees

def lookup_fee(subject, weekend=False):
    if not os.path.exists(FEE_FILE):
        return None
    fees = _load_fees(os.path.getmtime(FEE_FILE))
    s = subject.strip()
    if s in fees:
        return fees[s]
    if weekend and f"{s}/주말" in fees:
        return fees[f"{s}/주말"]
    return None

# ── GitHub 기반 DB 영구 저장 (Streamlit Cloud 로컬 디스크는 슬립/재배포 시 초기화될 수 있음) ──
GITHUB_DB_PATH = "backup_salesdb.db"  # 로컬 salesdb.db와 별도 경로 (로컬 gitignore와 안 겹치게)

def _github_cfg():
    if not hasattr(st, "secrets"):
        return None, None
    token = st.secrets.get("GITHUB_TOKEN", "")
    if not token:
        return None, None
    return token, st.secrets.get("GITHUB_REPO", "songseonggi22-boop/list")

def restore_db_from_github():
    token, repo = _github_cfg()
    if not token or os.path.exists(DB):
        return
    import base64, urllib.request, json as _json
    url = f"https://api.github.com/repos/{repo}/contents/{GITHUB_DB_PATH}"
    try:
        req = urllib.request.Request(url, headers={"Authorization": f"token {token}"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            content = base64.b64decode(_json.load(resp)["content"])
        with open(DB, "wb") as f:
            f.write(content)
    except Exception:
        pass  # 백업이 아직 없거나 네트워크 문제 → 새 DB로 시작

def backup_db_to_github():
    token, repo = _github_cfg()
    if not token:
        return
    import base64, urllib.request, json as _json
    url = f"https://api.github.com/repos/{repo}/contents/{GITHUB_DB_PATH}"
    headers = {"Authorization": f"token {token}"}
    sha = None
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=10) as resp:
            sha = _json.load(resp)["sha"]
    except Exception:
        pass
    try:
        with open(DB, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode()
        body = {"message": "auto: DB 백업", "content": content_b64}
        if sha:
            body["sha"] = sha
        req = urllib.request.Request(url, data=_json.dumps(body).encode(),
                                      headers={**headers, "Content-Type": "application/json"}, method="PUT")
        urllib.request.urlopen(req, timeout=10)
    except Exception:
        pass  # ponytail: 네트워크 실패는 조용히 무시, 다음 쓰기 때 다시 시도됨

# ── DB ───────────────────────────────────────────────────────
@st.cache_resource
def get_db():
    restore_db_from_github()
    c = sqlite3.connect(DB, check_same_thread=False)
    c.executescript("""
    CREATE TABLE IF NOT EXISTS tasks(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        category TEXT DEFAULT '주중업무',
        task_date TEXT DEFAULT '',
        assignee TEXT DEFAULT '',
        is_done INTEGER DEFAULT 0,
        created_at INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS consultations(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL, sched_date TEXT NOT NULL,
        sched_time TEXT DEFAULT '', expected_revenue INTEGER DEFAULT 0,
        ctype TEXT DEFAULT '단과', assignee TEXT DEFAULT '',
        created_at INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS daily_log(
        log_date TEXT PRIMARY KEY,
        team_name TEXT DEFAULT '2-2팀',
        rep1_name TEXT DEFAULT '', rep1_pct INTEGER DEFAULT 60,
        rep2_name TEXT DEFAULT '',
        rep1_call TEXT DEFAULT '', rep2_call TEXT DEFAULT '',
        done_count INTEGER DEFAULT 0,
        registered INTEGER DEFAULT 0, cod INTEGER DEFAULT 0, unregistered INTEGER DEFAULT 0,
        actual_revenue INTEGER DEFAULT 0, refund INTEGER DEFAULT 0,
        interview_count INTEGER DEFAULT 0,
        ddaz_num INTEGER DEFAULT 0, ddaz_den INTEGER DEFAULT 32,
        tmr_target INTEGER DEFAULT 0,
        month_target INTEGER DEFAULT 0, month_achieved INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS app_state(
        key TEXT PRIMARY KEY,
        value TEXT);

    CREATE TABLE IF NOT EXISTS week_target(
        month TEXT, week_no INTEGER, assignee TEXT DEFAULT '',
        start_date TEXT, end_date TEXT, target INTEGER DEFAULT 0,
        PRIMARY KEY (month, week_no, assignee));

    CREATE TABLE IF NOT EXISTS tt_upload(
        name TEXT PRIMARY KEY,
        content_b64 TEXT NOT NULL,
        uploaded_at INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS consultant(
        name TEXT PRIMARY KEY,
        phone TEXT DEFAULT '',
        mobile TEXT DEFAULT '',
        email TEXT DEFAULT '',
        sort INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS memo_tpl(
        key TEXT PRIMARY KEY,
        text TEXT NOT NULL,
        updated_at INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS lunch(
        person TEXT, ymd TEXT, amount INTEGER DEFAULT 0,
        PRIMARY KEY(person, ymd));

    CREATE TABLE IF NOT EXISTS lunch_day(
        ymd TEXT PRIMARY KEY, total INTEGER DEFAULT 0, paid INTEGER DEFAULT 0);

    CREATE TABLE IF NOT EXISTS lunch_paid(
        person TEXT, ymd TEXT, paid INTEGER DEFAULT 0,
        PRIMARY KEY(person, ymd));

    CREATE TABLE IF NOT EXISTS lunch_week_paid(
        person TEXT, monday TEXT, paid INTEGER DEFAULT 0,
        PRIMARY KEY(person, monday));
    """)
    c.commit()
    # 조직 개편: 2-3팀 → 2-2팀 (2026-08). 기존 행/GitHub 백업 DB에도 반영 (idempotent)
    c.execute("UPDATE daily_log SET team_name='2-2팀' WHERE team_name='2-3팀'")
    c.commit()
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN ctype TEXT DEFAULT '단과'")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN assignee TEXT DEFAULT ''")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN assignee TEXT DEFAULT ''")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN result_status TEXT DEFAULT ''")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN visit_type TEXT DEFAULT '신규방문'")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN actual_amount INTEGER DEFAULT 0")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN finalized INTEGER DEFAULT 0")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("ALTER TABLE consultations ADD COLUMN deposit_pct INTEGER DEFAULT 70")
        c.commit()
    except sqlite3.OperationalError:
        pass  # 이미 있음
    try:
        c.execute("SELECT assignee FROM week_target LIMIT 1")
    except sqlite3.OperationalError:
        # 기존 week_target은 (month,week_no)만 PK라 assignee 컬럼을 못 넣음 → 재생성
        c.execute("ALTER TABLE week_target RENAME TO week_target_old")
        c.execute("""CREATE TABLE week_target(
            month TEXT, week_no INTEGER, assignee TEXT DEFAULT '',
            start_date TEXT, end_date TEXT, target INTEGER DEFAULT 0,
            PRIMARY KEY (month, week_no, assignee))""")
        c.execute("""INSERT INTO week_target(month,week_no,assignee,start_date,end_date,target)
                     SELECT month,week_no,'',start_date,end_date,target FROM week_target_old""")
        c.execute("DROP TABLE week_target_old")
        c.commit()
    return c

# ponytail: global lock, connection pooling if concurrent traffic becomes a bottleneck
_db_lock = threading.Lock()

def q(sql, a=()):
    with _db_lock:
        return get_db().execute(sql, a).fetchall()

def run(sql, a=()):
    with _db_lock:
        get_db().execute(sql, a)
        get_db().commit()
    backup_db_to_github()  # ponytail: 매 쓰기마다 동기 백업이라 약간 느려질 수 있음, 부담되면 디바운스 추가

def get_state(key, default=None):
    rows = q("SELECT value FROM app_state WHERE key=?", (key,))
    return rows[0][0] if rows else default

def set_state(key, value):
    run("INSERT INTO app_state(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value))

# ── 공유 시간표 업로드 (팀 전체가 같은 시간표를 쓰도록 DB에 저장 → GitHub DB 백업으로 동기화) ──
def tt_uploads():
    return q("SELECT name, content_b64 FROM tt_upload ORDER BY name")

def save_tt_upload(name, raw_bytes):
    import base64
    run("INSERT INTO tt_upload(name,content_b64,uploaded_at) VALUES(?,?,?) "
        "ON CONFLICT(name) DO UPDATE SET content_b64=excluded.content_b64, uploaded_at=excluded.uploaded_at",
        (name, base64.b64encode(raw_bytes).decode(), int(time.time())))

def del_tt_upload(name):
    run("DELETE FROM tt_upload WHERE name=?", (name,))

def clear_tt_uploads():
    run("DELETE FROM tt_upload")

# ── 담당자(상담시간표/개강안내문/견적서 서명) 명단 · 팀 공유 ──
_CONSULT_SEED = dict(name="", phone="042-710-8921", mobile="010-2394-6693",
                     email="dkswo1215@koreaedugroup.com")

def get_consultants():
    rows = q("SELECT name, phone, mobile, email, sort FROM consultant ORDER BY sort, name")
    if not rows:
        run("INSERT INTO consultant(name,phone,mobile,email,sort) VALUES(?,?,?,?,0)",
            (_CONSULT_SEED["name"], _CONSULT_SEED["phone"], _CONSULT_SEED["mobile"], _CONSULT_SEED["email"]))
        rows = q("SELECT name, phone, mobile, email, sort FROM consultant ORDER BY sort, name")
    return [dict(name=n or "", phone=p or "", mobile=m or "", email=e or "", sort=s or 0)
            for n, p, m, e, s in rows]

def save_consultants(rows):
    """rows: list[dict] (name/phone/mobile/email). 전체 재작성 → 삭제도 반영. 이름 없으면 제외."""
    keep, seen = [], set()
    for r in rows:
        nm = str(r.get("name") or "").strip()
        if not nm or nm in seen:
            continue
        seen.add(nm)
        keep.append((nm, str(r.get("phone") or "").strip(), str(r.get("mobile") or "").strip(),
                     str(r.get("email") or "").strip(), len(keep)))
    if not keep:  # 전부 지우면 시드 1건 유지 (문서가 빈 담당자로 나가지 않게)
        keep = [(_CONSULT_SEED["name"], _CONSULT_SEED["phone"], _CONSULT_SEED["mobile"], _CONSULT_SEED["email"], 0)]
    with _db_lock:
        c = get_db()
        c.execute("DELETE FROM consultant")
        c.executemany("INSERT INTO consultant(name,phone,mobile,email,sort) VALUES(?,?,?,?,?)", keep)
        c.commit()
    backup_db_to_github()

# ── 메모 양식 (카테고리별 텍스트 양식 12종) · 팀 공유 편집 ──
# 원본: 클로드놀이_배포판/index.html MEMO_CATEGORIES / MEMO_DEFAULTS. 순서·전문 그대로.
MEMO_LABELS = [
    ("contactMemo", "컨택메모"),
    ("consultDone", "상담완료메모"),
    ("reservation", "예약메세지"),
    ("onlineClass", "온라인강의"),
    ("afterContact", "연결 안내메세지"),
    ("contactFail", "미연결 안내메세지"),
    ("bankInfo", "입금계좌안내"),
    ("refund", "환불일지"),
    ("remoteCode", "비대면 참여코드"),
    ("adobeSub", "어도비 학생구독"),
    ("midJoin", "중도투입"),
    ("depositReport", "입금 파악"),
]

MEMO_DEFAULTS = {
    "consultDone": "*상담자 : 여기에_본인_이름을_입력하세요\n*직업 :\n*나이 :\n*거주지 : 대전\n*과정 관심 계기 :\n*관심 및 희망 과정 :\n*수강 목적 :\n*학교&직장 내 특이사항 : -\n*평소 가용 가능한 시간 :\n*희망 개강 시기 :\n*희망 타임 :\n*수강 예산 :\n*성격 및 성향 : 좋음\n*결제권자/완납 예정 : 완납\n*환불 리스크 :\n*추가 등록 가망 과목/시기 :\n*이외 상담 내용 :\n\n\n\n**이번 상담 때 나의 반성 및 고칠 점 :",
    "contactMemo": "*컨텍자 : 여기에_본인_이름을_입력하세요\n*본인확인 및 문의여부 체크 : O\n*배워보시려는 계기 및 수강 희망 시기 :\n*분야 경험 및 수강 경험 여부 :\n*현재 거주지 및 평소 스케줄 어떤지 (가용시간체크) :\n*상담 명분 제시 및 상담 일정 확정 : 수업 구성 및 스케줄 안내\n*이외 상담내용 :\n\n\n\n** 수강료 / 시간표에 대한 나의 응대 : 따로 언급 x\n*같이 방문하실 지인 / 기타 특이사항 :\n* 통화시간 : 분",
    "midJoin": "상담일자 : 2026.00.00\n등록일자 : 2026.00.00\n요청자 / 해당 수업 : 여기에_본인_이름을_입력하세요 멘토 / 00.00 00시 포토샵\n교강사 :\n특이사항 :",
    "adobeSub": "어도비 학생 구독 안내\n\nhttps://www.adobe.com/kr/creativecloud/buy/students.html\n\n어도비 홈페이지 접속하신 후\n\n1) 왼쪽 상단 : [크리에이티비티 및 디자인] 클릭\n2) 구매 대상 : [학생 및 교사] 선택\n3) 플랜 또는 프로그램 선택 후 이메일 작성\n4) 교육 기관에 SBS아카데미컴퓨터아트학원 입력\n\n이후 구독 진행해주시면 됩니다!",
    "remoteCode": "■비대면 참여코드 안내■\n\nSBS아카데미AIX학원 대전입니다.\n\n과정 : 마야4\n개강 : 07.15\n종강 : 08.10\n시간 : 19:00-22:00\n접속코드 : https://discord.gg/bKB7R4JjrU\n\n과정 : AI비주얼디자인\n개강 : 07.22\n종강 : 08.04\n시간 : 13:00-15:00\n접속코드 : https://discord.gg/APwvkCRWhC\n\n과정 : AI마케팅영상제작\n개강 : 08.05\n종강 : 08.19\n시간 : 13:00-15:00\n접속코드 : https://discord.gg/6TQNmcT7mq",
    "refund": "-최초요청일: 26.00.00\n-면담일: 26.00.00\n-수강전/후:\n-담당멘토: 여기에_본인_이름을_입력하세요\n-등록과목:\n-등록금액:\n-공제금액:\n-실환불액:\n-결제수단:\n-중퇴사유:\n\n환불계좌 :\n\n\n\n환불보고\n\n-수강생:\n-요청시점:\n-어플접수:\n-등록과목:\n-공제내용:\n-등록금액:\n-공제금액:\n-실환불액:\n-요청내용:",
    "bankInfo": "입금계좌 안내(대전지점)\n\n오프라인\n입금은행 : 기업은행\n계좌번호 : 142-149362-01-021\n입금금액 : 0원\n예금주 : (주)에스씨에이아카데미대전\n\n입금계좌 안내(대전지점)\n\n온라인\n입금은행 : 기업은행\n계좌번호 : 336-103411-04-020\n입금금액 : 0원\n예금주 : (주)코리아온라인클래스\n\n현금영수증 발행시 따로 번호 알려 주시면 처리 도와드립니다.\n\n\n생년월일 :\n이메일 :\n거주지 : 동까지만 부탁드릴게요 !\n           ex) 대전 서구 둔산동",
    "contactFail": "안녕하세요 !\n\n홈페이지 통해서 컴활 자격증 과정으로 문의 남겨주신\n대전SBS아카데미AIX학원 여기에_본인_이름을_입력하세요 교육멘토입니다.\n\n문의 남겨주신 내용 관련해서 안내 도와드리고자\n042번호로 연락 드렸으나 통화 힘드신 것 같아\n카톡 남겨드려요 !\n\n혹시 통화 괜찮으신 시간대 있으실까요 ?",
    "afterContact": "안녕하세요 !\n\n조금 전에  과정 안내차 연락드린\nSBS아카데미 여기에_본인_이름을_입력하세요 교육멘토입니다.\n\n명함 남겨드리며 현재 학원에서 진행 중인 수업과\n온라인 강의까지 살펴보실 수 있게 링크 첨부드리니\n참고 해주시면 감사하겠습니다 !\n\n궁금하신 사항 있으시면 연락 부탁드릴게요 !\n\n📌 학원홈페이지 : http://sbsdjartcom.com/\n📌 온라인 강의 : https://www.ddazua.com/",
    "onlineClass": "📍 따즈아 온라인 강의 안내\n\n─────────────────\n📌 홈페이지\nhttps://www.ddazua.com/\n\n📌 쿠폰 코드\n\n─────────────────\n📍 쿠폰 코드 입력 방법\n\n1. 회원가입 후 로그인\n\n2. 마이페이지 ▶ 쿠폰 관리 이동\n\n3. 쿠폰 코드 입력 후 적용\n─────────────────\n📍 온라인 강의 이용 방법\n\n1. 내 강의실 이동\n\n2. 프리패스 강좌 선택\n\n3. 수강하실 강의 선택 후 학습 시작\n─────────────────\n\n추가로 궁금하신 사항이 있으시면\n언제든지 편하게 연락 주세요 😊",
    "reservation": "📍SBS아카데미AIX학원\n\n안녕하세요! 여기에_본인_이름을_입력하세요 멘토입니다 😊\n정확한 상담 도와드리기 위해 예약 안내드립니다.\n\n─────────────────\n[SBS아카데미 / 대전지점]\n📌 예약자명 : 000 님\n📌 교육과정 :\n📌 상담일정 : 00월 00일() 00:00\n─────────────────\n\n📍 주소\n대전 서구 둔산동 1110\n굿모닝어학원빌딩 9층\n\n🚗 주차 안내\n지하 기계식주차 가능\n※ 대형 SUV는 진입 불가하므로\n주변 공영주차장 이용 부탁드립니다.\n\n📞 문의전화\n042-710-8921\n\n⏰ 일정 변경이 필요하신 경우\n미리 연락 주시면 감사하겠습니다 :)\n\n─────────────────\nSBS아카데미AIX학원 대전지점\n멘토 여기에_본인_이름을_입력하세요 드림",
    # depositReport 는 원본이 계산 모달(15:00 보고)이라 양식 텍스트가 없어 스켈레톤으로 새로 작성 (특수 계산 폼은 범위 밖)
    "depositReport": "[15시 입금 파악]\n\n■ 건별 (시각 / 구분 / 방식 / 금액)\n- 00:00 신규 방문 000만원\n- 00:00 추가 온라인 000만원\n\n■ 익일 · 모레 · 따즈아\n- 익일 상담/예정 : 0건 / 0만원\n- 모레 예정 : 0건 / 0만원\n- 따즈아(온라인) : 0건 / 0만원\n\n■ 특이사항 :",
}

def get_memo_tpl(key):
    rows = q("SELECT text FROM memo_tpl WHERE key=?", (key,))
    return rows[0][0] if rows else MEMO_DEFAULTS.get(key, "")

def save_memo_tpl(key, text):
    run("INSERT INTO memo_tpl(key,text,updated_at) VALUES(?,?,?) "
        "ON CONFLICT(key) DO UPDATE SET text=excluded.text, updated_at=excluded.updated_at",
        (key, text, int(time.time())))

def reset_memo_tpl(key):
    run("DELETE FROM memo_tpl WHERE key=?", (key,))

def memo_fill(text, consultant):
    """표시·복사용 치환. consultant 값이 비어 있으면 해당 플레이스홀더는 그대로 둔다 (개강안내 방식)."""
    c = consultant or {}
    for ph, val in (("여기에_본인_이름을_입력하세요", (c.get("name") or "").strip()),
                    ("042-710-8921", (c.get("phone") or "").strip()),
                    ("010-2394-6693", (c.get("mobile") or "").strip()),
                    ("dkswo1215@koreaedugroup.com", (c.get("email") or "").strip())):
        if val:
            text = text.replace(ph, val)
    return text

# ── 점심값 정산 (팀 공유 · DB → GitHub 백업 동기화) ──
def lunch_members():
    return [x.strip() for x in get_state("lunch_members", "성기,윤호,연수").split(",") if x.strip()]

def set_lunch_members(names):
    set_state("lunch_members", ",".join(names))

def get_lunch_week(monday_iso, people):
    d0 = date.fromisoformat(monday_iso)
    days = [(d0 + timedelta(days=i)).isoformat() for i in range(7)]
    rows = q("SELECT person, ymd, amount FROM lunch WHERE ymd BETWEEN ? AND ?", (days[0], days[6]))
    m = {(p, y): a for p, y, a in rows}
    return days, {p: [int(m.get((p, dy), 0)) for dy in days] for p in people}

def save_lunch(person, ymd, amount):
    run("INSERT INTO lunch(person,ymd,amount) VALUES(?,?,?) "
        "ON CONFLICT(person,ymd) DO UPDATE SET amount=excluded.amount",
        (person, ymd, int(amount)))

def lunch_month_totals(ym):  # 'YYYY-MM' → {day(int): 합계}
    rows = q("SELECT ymd, SUM(amount) FROM lunch WHERE ymd LIKE ? GROUP BY ymd", (ym + "-%",))
    return {int(y[8:10]): int(t or 0) for y, t in rows}

# ── 날짜별 실제 결제금액 (사람별 금액 합계와 별개로 직접 입력) ──
def get_lunch_days(monday_iso):  # 그 주 7일 → {ymd: total(int)}
    d0 = date.fromisoformat(monday_iso)
    days = [(d0 + timedelta(days=i)).isoformat() for i in range(7)]
    rows = q("SELECT ymd, total FROM lunch_day WHERE ymd BETWEEN ? AND ?", (days[0], days[6]))
    m = {y: int(t or 0) for y, t in rows}
    return {dy: m.get(dy, 0) for dy in days}

def save_lunch_day(ymd, total):  # lunch_day.paid 컬럼은 미사용 (담당자별 lunch_paid 로 대체)
    run("INSERT INTO lunch_day(ymd,total) VALUES(?,?) "
        "ON CONFLICT(ymd) DO UPDATE SET total=excluded.total",
        (ymd, int(total)))

def lunch_month_days(ym):  # 'YYYY-MM' → {day(int): total(int)}
    rows = q("SELECT ymd, total FROM lunch_day WHERE ymd LIKE ?", (ym + "-%",))
    return {int(y[8:10]): int(t or 0) for y, t in rows}

# ── 담당자별 냈음: 사람×날짜(그날 자기 몫) + 사람×주차(정산액 입금) ──
def get_lunch_paid(monday_iso, people):  # {(person, ymd): bool}
    d0 = date.fromisoformat(monday_iso)
    days = [(d0 + timedelta(days=i)).isoformat() for i in range(7)]
    rows = q("SELECT person, ymd, paid FROM lunch_paid WHERE ymd BETWEEN ? AND ?", (days[0], days[6]))
    m = {(p, y): bool(v) for p, y, v in rows}
    return {(p, dy): m.get((p, dy), False) for p in people for dy in days}

def save_lunch_paid(person, ymd, paid):
    run("INSERT INTO lunch_paid(person,ymd,paid) VALUES(?,?,?) "
        "ON CONFLICT(person,ymd) DO UPDATE SET paid=excluded.paid",
        (person, ymd, 1 if paid else 0))

def get_lunch_week_paid(monday_iso, people):  # {person: bool}
    rows = q("SELECT person, paid FROM lunch_week_paid WHERE monday=?", (monday_iso,))
    m = {p: bool(v) for p, v in rows}
    return {p: m.get(p, False) for p in people}

def save_lunch_week_paid(person, monday, paid):
    run("INSERT INTO lunch_week_paid(person,monday,paid) VALUES(?,?,?) "
        "ON CONFLICT(person,monday) DO UPDATE SET paid=excluded.paid",
        (person, monday, 1 if paid else 0))

def lunch_month_paid(ym, people):  # {day(int): 'full'|'partial'} — 그날 몫 있는 사람이 다 냈나
    owe, got = {}, {}
    for y, p in q("SELECT ymd, person FROM lunch WHERE ymd LIKE ? AND amount>0", (ym + "-%",)):
        if p in people:
            owe.setdefault(int(y[8:10]), set()).add(p)
    for y, p in q("SELECT ymd, person FROM lunch_paid WHERE ymd LIKE ? AND paid=1", (ym + "-%",)):
        if p in people:
            got.setdefault(int(y[8:10]), set()).add(p)
    out = {}
    for d in set(owe) | set(got):
        o, g = owe.get(d, set()), got.get(d, set())
        if o and o <= g:
            out[d] = "full"
        elif g:
            out[d] = "partial"
    return out

def get_team_members():
    try:
        return json.loads(get_state("team_members", "[]"))
    except Exception:
        return []

def add_team_member(name):
    members = get_team_members()
    if name and name not in members:
        members.append(name)
        set_state("team_members", json.dumps(members, ensure_ascii=False))

def remove_team_member(name):
    members = get_team_members()
    if name in members:
        members.remove(name)
        set_state("team_members", json.dumps(members, ensure_ascii=False))

def get_weekday_pct(person, day):
    return int(get_state(f"weekday_pct_{person}_{day}", "0") or 0)

def set_weekday_pct(person, day, pct):
    set_state(f"weekday_pct_{person}_{day}", str(int(pct)))

def get_tasks():
    rows = q("SELECT id,title,category,task_date,assignee,is_done FROM tasks ORDER BY is_done,created_at DESC")
    return [dict(zip("id title category task_date assignee is_done".split(), r)) for r in rows]

def get_consults():
    rows = q("""SELECT id,name,sched_date,sched_time,expected_revenue,ctype,assignee,result_status,visit_type,actual_amount,finalized,deposit_pct
                FROM consultations ORDER BY sched_date,sched_time""")
    return [dict(zip("id name sched_date sched_time expected_revenue ctype assignee result_status visit_type actual_amount finalized deposit_pct".split(), r))
            for r in rows]

def set_consult_status(cid, status):
    run("UPDATE consultations SET result_status=? WHERE id=?", (status, cid))

def set_consult_amount(cid, amount):
    run("UPDATE consultations SET actual_amount=? WHERE id=?", (amount, cid))

def set_consult_pct(cid, pct):
    run("UPDATE consultations SET deposit_pct=? WHERE id=?", (pct, cid))

def finalize_consult(cid):
    run("UPDATE consultations SET finalized=1 WHERE id=?", (cid,))

def unfinalize_consult(cid):
    run("UPDATE consultations SET finalized=0 WHERE id=?", (cid,))

def add_task(title, cat, d, assignee):
    run("INSERT INTO tasks(title,category,task_date,assignee,is_done,created_at) VALUES(?,?,?,?,0,?)",
        (title, cat, d, assignee, int(time.time()*1000)))

def toggle_task(tid, v): run("UPDATE tasks SET is_done=? WHERE id=?", (int(v), tid))
def del_task(tid):       run("DELETE FROM tasks WHERE id=?", (tid,))
def update_task(tid, title, cat, d, assignee):
    run("UPDATE tasks SET title=?, category=?, task_date=?, assignee=? WHERE id=?",
        (title, cat, d, assignee, tid))

def add_consult(name, d, t, rev, ctype, assignee, visit_type="신규방문"):
    run("""INSERT INTO consultations(name,sched_date,sched_time,expected_revenue,ctype,assignee,visit_type,created_at)
           VALUES(?,?,?,?,?,?,?,?)""",
        (name, d, t, rev, ctype, assignee, visit_type, int(time.time()*1000)))
def del_consult(cid): run("DELETE FROM consultations WHERE id=?", (cid,))

def update_consult(cid, name, d, t, rev, ctype, assignee, visit_type="신규방문"):
    run("""UPDATE consultations SET name=?, sched_date=?, sched_time=?, expected_revenue=?, ctype=?, assignee=?, visit_type=?
           WHERE id=?""", (name, d, t, rev, ctype, assignee, visit_type, cid))

# ── 일지(daily_log): 자동계산 외 수동 항목 저장 ────────────────────
LOG_COLS = ("log_date team_name rep1_name rep1_pct rep2_name rep1_call rep2_call "
            "done_count registered cod unregistered actual_revenue refund "
            "interview_count ddaz_num ddaz_den tmr_target month_target month_achieved").split()

def get_log(d):
    rows = q(f"SELECT {','.join(LOG_COLS)} FROM daily_log WHERE log_date=?", (d,))
    if rows:
        return dict(zip(LOG_COLS, rows[0]))
    prev = q(f"SELECT {','.join(LOG_COLS)} FROM daily_log ORDER BY log_date DESC LIMIT 1")
    base = dict(zip(LOG_COLS, prev[0])) if prev else {}
    new_row = dict(log_date=d, team_name=base.get("team_name", "2-2팀"),
                    rep1_name=base.get("rep1_name", ""), rep1_pct=base.get("rep1_pct", 60),
                    rep2_name=base.get("rep2_name", ""), rep1_call="", rep2_call="",
                    done_count=0, registered=0, cod=0, unregistered=0, actual_revenue=0, refund=0,
                    interview_count=0, ddaz_num=0, ddaz_den=base.get("ddaz_den", 32),
                    tmr_target=0, month_target=base.get("month_target", 0), month_achieved=0)
    run("""INSERT INTO daily_log(log_date, team_name, rep1_name, rep1_pct, rep2_name, ddaz_den, month_target)
           VALUES(?,?,?,?,?,?,?)""",
        (d, new_row["team_name"], new_row["rep1_name"], new_row["rep1_pct"],
         new_row["rep2_name"], new_row["ddaz_den"], new_row["month_target"]))
    return new_row

# ── 구글 시트 백업 (주기적 백업용, DB는 여전히 sqlite) ───────────
SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

@st.cache_resource
def get_sheets_client():
    if not hasattr(st, "secrets") or "gcp_service_account" not in st.secrets:
        return None
    from google.oauth2.service_account import Credentials
    import gspread
    creds = Credentials.from_service_account_info(dict(st.secrets["gcp_service_account"]), scopes=SHEETS_SCOPES)
    return gspread.authorize(creds)

def backup_to_sheets():
    import gspread
    client = get_sheets_client()
    sheet_id = st.secrets.get("SHEETS_SPREADSHEET_ID", "") if hasattr(st, "secrets") else ""
    if not client or not sheet_id:
        raise RuntimeError("구글 시트 백업 설정이 안 되어 있어요 (Secrets에 gcp_service_account / SHEETS_SPREADSHEET_ID 확인).")
    sh = client.open_by_key(sheet_id)

    def write_sheet(name, headers, rows):
        try:
            ws = sh.worksheet(name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=name, rows=max(100, len(rows) + 10), cols=max(10, len(headers)))
        ws.clear()
        ws.update([headers] + [[str(v) for v in row] for row in rows])

    tasks_ = get_tasks()
    write_sheet("tasks", ["id", "title", "category", "task_date", "assignee", "is_done"],
                [[t["id"], t["title"], t["category"], t["task_date"], t["assignee"], t["is_done"]] for t in tasks_])

    consults_ = get_consults()
    write_sheet("consultations", ["id", "name", "sched_date", "sched_time", "expected_revenue", "ctype"],
                [[c["id"], c["name"], c["sched_date"], c["sched_time"], c["expected_revenue"], c["ctype"]] for c in consults_])

    logs_ = q(f"SELECT {','.join(LOG_COLS)} FROM daily_log ORDER BY log_date")
    write_sheet("daily_log", LOG_COLS, [list(r) for r in logs_])

# ── 금액 입력 (콤마 표시, number_input의 스크롤휠 오작동 방지용 텍스트 입력) ──
def money_input(label, value, key, container=None, **kw):
    box = container if container is not None else st
    raw = box.text_input(label, value=f"{int(value):,}", key=key, **kw)
    digits = re.sub(r"[^\d]", "", raw)
    return int(digits) if digits else 0

# ── 15시보고/마감보고 편집 가능한 양식 ({토큰} 치환) ────────────────
PM3_TEMPLATE_DEFAULT = """[{팀명} 15:00 보고]
{입금완료} / {입금예정}
상담중 {상담중}
{상담목록}
익일상담 {익일상담} / 익일예정 {익일예정}
모레상담 {모레상담} / 모레예정 {모레예정}
익일면접 {익일면접}건
따즈아 {따즈아분자} / {따즈아분모}"""

CLOSE_TEMPLATE_DEFAULT = """컴퓨터 {팀명} 영업마감보고

 상담 : {상담건수}
 등록 : {등록}
 COD : {COD}
 미등록 : {미등록}

금일매출결과 :{금일매출}원
환불 :{환불}원

통화시간
{담당자1} {통화1} (상담 {상담건수}건)
{담당자2} {통화2}

익일예정상담 : {익일상담}건
익일예정매출 : {익일매출}원
익일목표매출 : {익일목표매출}원

{월}월 팀목표매출 : {팀목표매출}
현재달성매출 : {현재달성매출}
현재달성율 : {달성율}%"""

def render_report_template(template, values):
    try:
        return template.format(**values)
    except (KeyError, IndexError) as e:
        return f"[양식 오류: {e} 토큰을 찾을 수 없어요. 아래 사용 가능한 토큰 목록을 확인해주세요]\n\n" + template

def report_template_editor(label, key, default_tmpl, tokens):
    with st.expander(f"✏️ {label} 양식 편집"):
        st.caption("사용 가능한 토큰: " + " ".join(f"{{{t}}}" for t in tokens))
        cur = get_state(key, default_tmpl)
        edited = st.text_area(f"{label} 양식", value=cur, height=200, key=f"{key}_edit", label_visibility="collapsed")
        c1, c2 = st.columns(2)
        if c1.button("💾 저장", key=f"{key}_save", use_container_width=True):
            set_state(key, edited)
            st.rerun()
        if c2.button("↩️ 기본값으로 되돌리기", key=f"{key}_reset", use_container_width=True):
            set_state(key, default_tmpl)
            st.rerun()
    return get_state(key, default_tmpl)

# ── 배정 문구 생성기 ──────────────────────────────────────────
def fdate(iso):
    if not iso: return ""
    p = iso.split("-")
    return f"{int(p[1])}.{int(p[2])}"

def gen_text(type_, **k):
    nd, nt = fdate(k.get("nd","")), k.get("nt","")
    ns, nts = k.get("ns",""), k.get("nts","")
    ns_full = f"{ns}/{nts}"
    nfee, ofee = k.get("nfee",0), k.get("ofee",0)
    od, ot, os_ = fdate(k.get("od","")), k.get("ot",""), k.get("os_","")

    if type_ == "신규":    return f"미배정 -> {nd} {nt} {ns_full} 배정"
    if type_ == "과목변경":
        if ofee or nfee:
            diff = abs(nfee - ofee)
            sfx  = f"\n[차액 {diff:,}원 무시]" if diff else ""
            return f"미배정 {os_} (수강료:{ofee:,}원) -> {nd} {nt} {ns_full} (수강료:{nfee:,}원) 과목변경 배정{sfx}"
        return f"미배정 {os_} -> {nd} {nt} {ns_full} 과목변경 배정"
    if type_ == "취소":    return f"{od} {ot} {os_} 배정 -> 미배정"
    if type_ == "날짜변경": return f"{od} {ot} {os_} 배정 -> {nd} {nt} {os_} 배정"

# ── 페이지 설정 ───────────────────────────────────────────────
st.set_page_config(page_title="업무 대시보드", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

ss = st.session_state
if "assign_out" not in ss: ss.assign_out = ""
if "notice_out" not in ss: ss.notice_out = ""

now_kst   = datetime.now(KST)
today     = now_kst.date()
today_str = today.isoformat()
yr, mo    = today.year, today.month

# 한국시간 06:00 기준으로 하루 업무일 판단 → 새 업무일이면 체크 초기화
workday = (today - timedelta(days=1)).isoformat() if now_kst.hour < 6 else today_str
if get_state("last_task_reset") != workday:
    run("UPDATE tasks SET is_done=0")
    set_state("last_task_reset", workday)

# 업무일 기준 이번 주 월요일 → 주중업무 게이지는 월요일이 되면 새로 시작
_workday_date = date.fromisoformat(workday)
week_start = (_workday_date - timedelta(days=_workday_date.weekday())).isoformat()

# 하루 한 번 자동 구글시트 백업 (설정 안 돼있으면 조용히 건너뜀)
if get_state("last_sheets_backup") != today_str and get_sheets_client():
    try:
        backup_to_sheets()
        set_state("last_sheets_backup", today_str)
    except Exception:
        pass  # 다음 로드 때 다시 시도

# ── 데이터 로드 ────────────────────────────────────────────────
tasks    = get_tasks()
consults = get_consults()
today_c  = [c for c in consults if c["sched_date"] == today_str]

tomorrow_str = (today + timedelta(days=1)).isoformat()
dayafter_str = (today + timedelta(days=2)).isoformat()

def day_stats(d):
    items = [c for c in consults if c["sched_date"] == d]
    cnt = len(items)
    rev = sum(c["expected_revenue"] for c in items)
    reg = sum(1 for c in items if c["ctype"] == "정규")
    return cnt, rev, reg, cnt - reg

today_cnt, today_rev, today_reg, today_dan = day_stats(today_str)
tmr_cnt, tmr_rev, _, _ = day_stats(tomorrow_str)
daf_cnt, daf_rev, _, _ = day_stats(dayafter_str)
log = get_log(today_str)

# 달력용 이벤트 맵 (완료 처리된 상담은 캘린더에서 제외)
cmap = {}
for c in consults:
    if c["finalized"]:
        continue
    if c["sched_date"][:7] == f"{yr:04d}-{mo:02d}":
        d = int(c["sched_date"][8:])
        cmap.setdefault(d, []).append(c)

# ── CSS (디자인 시스템: 클로드놀이_배포판 블루 톤 이식 · Phase 1) ──────────────
# 기준: AI 위키 wiki/design/업무대시보드-디자인시스템.md
st.markdown("""<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css');
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');

#MainMenu,footer,[data-testid="stHeader"],[data-testid="stToolbar"]{display:none!important}
.block-container{padding:.7rem 1.8rem 2.5rem!important;max-width:100%!important}
[data-testid="stAppViewContainer"],.main{background:var(--color-bg)!important}

:root{
  /* 색상 토큰 (블루 톤) */
  --color-bg:#F5F7F9; --color-panel:#FFFFFF; --color-surface:#FFFFFF; --color-border:#E7EBEF;
  --color-primary:#5C7A94; --color-primary-hover:#6D8AA3; --color-selected:#F1F4F6; --color-sidebar-active:#F1F4F6;
  --color-text:#333D46; --color-text-secondary:#707F8D; --color-muted:#9FACB7;
  --color-success:#8DA57A; --color-warning:#C79A68; --color-danger:#C68A7A; --color-urgent:#DD5C33;
  --color-written-bg:#EFEFF0; --color-written-text:#6B6F76; --color-practical-text:#5C7A94;
  /* 간격 / 반경 / 전이 스케일 */
  --space-1:4px;--space-2:6px;--space-3:8px;--space-4:10px;--space-5:12px;
  --space-6:16px;--space-7:20px;--space-8:24px;--space-9:32px;--space-10:40px;
  --radius-sm:8px;--radius-md:10px;--radius-lg:12px;--radius-pill:999px;
  --ease:.2s ease;
  /* 기존 클래스 호환 별칭 (점진 이식용) */
  --bg:var(--color-bg); --card:var(--color-surface); --tx:var(--color-text);
  --sub:var(--color-text-secondary); --bd:var(--color-border);
  --p1:#FDECEC; --p1t:#B5786A;   /* 담당자1 / 급함 계열 */
  --p2:#FAF2E6; --p2t:#B08A5A;   /* 경고 계열 */
  --p3:#EDF2F6; --p3t:#5C7A94;   /* primary 계열 */
}
*{font-family:'Pretendard','Noto Sans KR','Apple SD Gothic Neo','Malgun Gothic',sans-serif}
body{color:var(--color-text)}

@keyframes fadeInUp{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}

/* Streamlit 버튼 — 기본은 아웃라인, type="primary"는 슬레이트 채움 */
[data-testid="stButton"]>button{
  padding:4px 12px!important;font-size:11.5px!important;border-radius:var(--radius-sm)!important;
  font-weight:500!important;min-height:0!important;line-height:1.5!important;
  border:1px solid var(--color-border)!important;background:var(--color-surface)!important;color:var(--color-text-secondary)!important;
  transition:background var(--ease),color var(--ease),border-color var(--ease)!important}
[data-testid="stButton"]>button:hover{background:var(--color-sidebar-active)!important;color:var(--color-text)!important;border-color:var(--color-primary)!important}
[data-testid="stButton"]>button[kind="primary"]{
  background:var(--color-primary)!important;border-color:var(--color-primary)!important;color:#fff!important}
[data-testid="stButton"]>button[kind="primary"]:hover{background:var(--color-primary-hover)!important;border-color:var(--color-primary-hover)!important}
[data-testid="stCheckbox"] label{font-size:11.5px!important;color:var(--color-text-secondary)!important}
[data-testid="stCheckbox"]>label>div:first-child{width:13px!important;height:13px!important;border-radius:4px!important}
[data-testid="stCheckbox"] [data-baseweb="checkbox"] [data-checked="true"]{background:var(--color-primary)!important;border-color:var(--color-primary)!important}
[data-testid="stForm"]{border:none!important;padding:0!important}
.stTextInput input,.stSelectbox>div>div,.stDateInput input,.stNumberInput input,.stTextArea textarea{
  border-radius:var(--radius-md)!important;border:1.5px solid var(--color-border)!important;font-size:12px!important;background:var(--color-surface)!important;transition:border-color var(--ease)!important}
.stTextInput input:focus,.stDateInput input:focus,.stNumberInput input:focus,.stTextArea textarea:focus{border-color:var(--color-primary)!important;outline:none!important}
[data-testid="stExpander"]{border:1.5px solid var(--color-border)!important;border-radius:var(--radius-lg)!important;background:var(--color-surface)!important}
[data-testid="stTabs"] button[aria-selected="true"]{color:var(--color-primary)!important}
[data-testid="stTabs"] [data-baseweb="tab-highlight"]{background:var(--color-primary)!important}
div[data-testid="stHorizontalBlock"]{gap:.8rem!important}
hr{margin:.3rem 0!important;border-color:var(--color-border)!important}
::selection{background:var(--color-selected)}

/* 공통 카드 */
.db-card{background:var(--color-surface);border-radius:var(--radius-lg);padding:var(--space-8);
         box-shadow:0 1px 2px rgba(20,40,60,.04),0 8px 24px rgba(20,40,60,.05);border:1px solid var(--color-border);animation:fadeInUp .35s ease both}
.db-card-title{font-size:15px;font-weight:600;color:var(--color-text);
               display:flex;align-items:center;gap:7px;margin-bottom:var(--space-6)}
/* 우측 컬럼 = 2차 정보: 카드 제목 톤 다운 (본문 최상위 2컬럼만) */
div[data-testid="stHorizontalBlock"]>div[data-testid="stColumn"]:last-of-type .db-card-title{font-size:13.5px;font-weight:600;color:var(--color-text-secondary)}

/* 달력 (1차 정보 — 크게, 대비 높게) */
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);
          border-top:1px solid var(--color-border);border-left:1px solid var(--color-border);
          border-radius:0 0 var(--radius-sm) var(--radius-sm);overflow:hidden}
.cal-head{background:var(--color-panel);padding:8px 4px;text-align:center;font-weight:500;
          font-size:11.5px;text-transform:uppercase;letter-spacing:.04em;border-right:1px solid var(--color-border);border-bottom:1px solid var(--color-border);color:var(--color-muted)}
.cal-day{min-height:118px;padding:8px;border-right:1px solid var(--color-border);
         border-bottom:1px solid var(--color-border);background:var(--color-surface);vertical-align:top}
.cal-day.td{background:var(--color-selected)}
.cal-num{font-size:12px;font-weight:500;margin-bottom:3px;display:inline-block;
         width:21px;height:21px;line-height:21px;text-align:center;border-radius:50%}
.cal-num.td{background:var(--color-primary);color:#fff!important;font-weight:600}
.ci{font-size:9px;padding:2px 5px;border-radius:4px;margin-bottom:2px;
    white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-weight:600;display:block}
.ci1{background:var(--p1);color:var(--p1t)}.ci2{background:var(--p2);color:var(--p2t)}.ci3{background:var(--p3);color:var(--p3t)}

/* 칸반 (2차 정보) */
.k-head{font-size:12px;font-weight:600;padding:5px 11px;border-radius:var(--radius-sm);display:inline-block;margin-bottom:10px}
.kp1{background:var(--p1);color:var(--p1t)}.kp2{background:var(--p2);color:var(--p2t)}.kp3{background:var(--p3);color:var(--p3t)}
.t-card{background:var(--color-surface);border:1px solid var(--color-border);border-radius:var(--radius-md);
        padding:12px;margin-bottom:4px;box-shadow:0 1px 2px rgba(20,40,60,.04);
        transition:transform .15s var(--ease),box-shadow .15s var(--ease)}
.t-card:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(20,40,60,.08)}
.t-date{font-size:11px;color:var(--color-text-secondary);margin-bottom:4px}
.t-title{font-size:13px;font-weight:500;line-height:1.45;color:var(--color-text)}
.t-title.done{text-decoration:line-through;color:var(--color-muted)}

/* ── 좌측 사이드바 nav (클로드놀이_배포판 앱셸 톤) ── */
section[data-testid="stSidebar"]{background:var(--color-panel);border-right:1px solid var(--color-border);width:236px!important;min-width:236px!important}
section[data-testid="stSidebar"] .block-container{padding:1.4rem 1rem 2rem!important}
section[data-testid="stSidebar"] [role="radiogroup"]{gap:2px!important}
section[data-testid="stSidebar"] [role="radiogroup"] label{
  padding:8px 10px!important;border-radius:var(--radius-sm)!important;margin:0!important;
  font-size:13px!important;font-weight:500!important;color:var(--color-text-secondary)!important;
  transition:background var(--ease),color var(--ease)!important;cursor:pointer!important}
section[data-testid="stSidebar"] [role="radiogroup"] label:hover{background:var(--color-sidebar-active)!important;color:var(--color-text)!important}
section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked){background:var(--color-sidebar-active)!important;color:var(--color-primary)!important;font-weight:600!important}
section[data-testid="stSidebar"] [role="radiogroup"] [data-baseweb="radio"]>div:first-child{transform:scale(.72);opacity:.55}  /* 라디오 점 작게 */
.sb-brand{font-size:15px;font-weight:700;color:var(--color-text);padding:2px 6px 2px}
.sb-brand-sub{font-size:10.5px;color:var(--color-muted);padding:0 6px 14px}
.sb-sec{font-size:10px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;color:var(--color-muted);margin:16px 6px 4px}
</style>""", unsafe_allow_html=True)

def _course_picker(label, key):
    months = available_months()
    if not months:
        st.caption("시간표 데이터가 없어요.")
        return None

    sel_key = f"{key}_sel"
    saved = st.session_state.get(sel_key)  # (month, weekend, idx) or None

    month_tabs = st.tabs(months)
    for month, tab in zip(months, month_tabs):
        with tab:
            wd_choice = st.radio("구분", ["평일", "주말"], horizontal=True,
                                  key=f"{key}_{month}_wd", label_visibility="collapsed")
            weekend = wd_choice == "주말"
            sessions = sessions_in_month(month, weekend)
            if not sessions:
                st.caption("해당 달에 이 구분의 강좌가 없어요.")
                continue

            rooms = sorted({s["room"] for s in sessions})
            times = sorted({s["start_time"] for s in sessions})
            by_cell = {}
            for i, s in enumerate(sessions):
                by_cell.setdefault((s["room"], s["start_time"]), []).append((i, s))

            cur_idx = saved[2] if saved and saved[0] == month and saved[1] == weekend else None

            ROOMS_PER_ROW = 6
            for start in range(0, len(rooms), ROOMS_PER_ROW):
                row_rooms = rooms[start:start + ROOMS_PER_ROW]
                head = st.columns(len(row_rooms))
                for col, room in zip(head, row_rooms):
                    col.markdown(f"""<div style='font-size:11px;font-weight:700;text-align:center;
                                background:#EDF2F6;color:#5C7A94;border-radius:6px;padding:4px;margin-bottom:4px'>
                                {room}</div>""", unsafe_allow_html=True)
                for t in times:
                    if not any((room, t) in by_cell for room in row_rooms):
                        continue
                    row = st.columns(len(row_rooms))
                    for col, room in zip(row, row_rooms):
                        with col:
                            for i, s in by_cell.get((room, t), []):
                                is_sel = cur_idx == i
                                btn_label = (("✅ " if is_sel else "") +
                                             f"{t} {s['subject']} ({s['start_date'][5:]}~{s['end_date'][5:]})")
                                if st.button(btn_label, key=f"{key}_{month}_{weekend}_cell_{i}",
                                             use_container_width=True):
                                    st.session_state[sel_key] = (month, weekend, i)
                                    st.rerun()

    if saved:
        s_month, s_weekend, s_idx = saved
        sessions = sessions_in_month(s_month, s_weekend)
        if s_idx < len(sessions):
            chosen = sessions[s_idx]
            st.caption(f"선택됨: {chosen['subject']} ({chosen['room']}, {chosen['start_time']}~{chosen['end_time']}, "
                       f"{chosen['teacher']}, 개강 {chosen['start_date']})")
            return chosen
    return None

def render_assign_automation():
    # ── 시간표 배정 자동화 ─────────────────────────────────────────
    st.markdown("""
    <div class="db-card">
      <div class="db-card-title">📋 시간표 배정 자동화</div>
      <div style='font-size:12px;color:var(--color-muted);margin-top:-10px;margin-bottom:14px'>
        정보 입력 → 배정 문구 자동 생성 후 복사
      </div>
    </div>""", unsafe_allow_html=True)

    candidates = []  # 체크박스로 선택된 강좌 → 아래 "선택한 강좌 배정 문구 생성"에서 일괄 처리
    if "cb_gen_seq" not in st.session_state:
        st.session_state["cb_gen_seq"] = 0
    cb_seq = st.session_state["cb_gen_seq"]  # 생성 후 체크박스 초기화(재마운트)용

    with st.expander("🔍 강의장×시간 그리드에서 강좌 선택 (신규 배정용)", expanded=False):
        grid_months = available_months()
        if not grid_months:
            st.caption("시간표 데이터가 없어요.")
        else:
            grid_month_tabs = st.tabs(grid_months)
            for gmonth, gtab in zip(grid_months, grid_month_tabs):
                with gtab:
                    gwd_choice = st.radio("구분", ["평일", "주말"], horizontal=True,
                                           key=f"ttpick_{gmonth}_wd", label_visibility="collapsed")
                    gweekend = gwd_choice == "주말"
                    month_sessions = sessions_in_month(gmonth, gweekend)
                    if not month_sessions:
                        st.caption("해당 달에 이 구분의 강좌가 없어요.")
                        continue

                    rooms = sorted({s["room"] for s in month_sessions})
                    times = sorted({s["start_time"] for s in month_sessions})
                    by_cell = {}
                    for s in month_sessions:
                        by_cell.setdefault((s["room"], s["start_time"]), []).append(s)

                    ROOMS_PER_ROW = 6
                    for start in range(0, len(rooms), ROOMS_PER_ROW):
                        row_rooms = rooms[start:start + ROOMS_PER_ROW]
                        head_cols = st.columns(len(row_rooms))
                        for col, room in zip(head_cols, row_rooms):
                            col.markdown(f"""<div style='font-size:11px;font-weight:700;text-align:center;
                                        background:#EDF2F6;color:#5C7A94;border-radius:6px;padding:4px;margin-bottom:4px'>
                                        {room}</div>""", unsafe_allow_html=True)
                        for t in times:
                            if not any((room, t) in by_cell for room in row_rooms):
                                continue
                            row_cols = st.columns(len(row_rooms))
                            for col, room in zip(row_cols, row_rooms):
                                with col:
                                    for si, s in enumerate(by_cell.get((room, t), [])):
                                        weekend = _is_weekend_days(s["days"])
                                        label = (f"{t} {s['subject']} ({s['start_date'][5:]}~{s['end_date'][5:]})"
                                                 + (" [주말]" if weekend else ""))
                                        key = f"ttpick_{gmonth}_{gweekend}_{room}_{t}_{si}_{cb_seq}"
                                        st.checkbox(label, key=key)
                                        candidates.append(dict(key=key, nd=s["start_date"],
                                                                nt=s["start_time"], ns=s["subject"],
                                                                nts="주말" if weekend else ""))

    with st.expander("🔎 과목명으로 검색 (신규 배정용)"):
        search_q = st.text_input("과목명/강사명 검색", key="tt_search_q")
        if search_q:
            hits = [s for s in get_timetable()
                    if search_q.strip() in s["subject"] or search_q.strip() in s["teacher"]]
            hits = sorted(hits, key=lambda s: (s["subject"], s["start_time"]))
            if hits:
                for i, s in enumerate(hits):
                    weekend = _is_weekend_days(s["days"])
                    label = (f"{s['subject']} — {s['start_time']} ({s['room']}, {s['teacher']}) "
                             f"개강 {s['start_date']}~종강 {s['end_date']}") + ("  [주말]" if weekend else "")
                    key = f"ttsearch_{i}_{cb_seq}"
                    st.checkbox(label, key=key)
                    candidates.append(dict(key=key, nd=s["start_date"],
                                            nt=s["start_time"], ns=s["subject"], nts="주말" if weekend else ""))
            else:
                st.caption("일치하는 강좌가 없어요.")

    with st.expander("📎 개인 시간표 업로드해서 강좌 선택 (신규 배정용)"):
        upl = st.file_uploader("개인 시간표(xlsx)", type=["xlsx"], key="personal_tt_upl")
        if upl:
            try:
                student, courses = parse_personal_timetable(upl)
            except Exception as e:
                student, courses = "", []
                st.error(f"파일을 읽지 못했어요: {e}")
            if student:
                st.caption(f"{student}님의 시간표에서 {len(courses)}개 강좌를 찾았어요.")
            if courses:
                for i, c in enumerate(courses):
                    matched = match_real_session(c["subject"], c["start_date"])
                    if matched:
                        weekend = _is_weekend_days(matched["days"])
                        label = (f"{matched['subject']} — {matched['start_time']} ({matched['room']}, {matched['teacher']}) "
                                 f"개강 {matched['start_date']}" + ("  [주말]" if weekend else "") + "  ✅ 시간표에서 찾음")
                        nd, nt, ns = matched["start_date"], matched["start_time"], matched["subject"]
                    else:
                        weekend = _is_weekend_days(c["days"])
                        label = (f"{c['subject']} — {c['start_time']} 개강 {c['start_date']}"
                                 + ("  [주말]" if weekend else "") + "  ⚠️ 시간표에서 못 찾음(입력값 그대로 사용)")
                        nd, nt, ns = c["start_date"], c["start_time"], c["subject"]
                    key = f"pt_pick_{i}_{cb_seq}"
                    st.checkbox(label, key=key)
                    candidates.append(dict(key=key, nd=nd, nt=nt, ns=ns, nts="주말" if weekend else ""))
            elif upl and not student:
                st.caption("강좌 정보를 찾지 못했어요. 파일 형식을 확인해주세요.")

    with st.expander("🖼️ 시간표 이미지 업로드 (AI 인식, 실험적)"):
        st.caption("AI가 이미지를 읽어서 추출하는 거라 100% 정확하진 않아요. 채워진 값을 확인하고 쓰세요.")

        from streamlit_paste_button import paste_image_button
        if "tt_img_seq" not in st.session_state:
            st.session_state["tt_img_seq"] = 0
        seq = st.session_state["tt_img_seq"]

        paste_result = paste_image_button("📋 클립보드에서 붙여넣기", key=f"tt_img_paste_{seq}")
        img = st.file_uploader("또는 파일로 업로드", type=["png", "jpg", "jpeg"], key=f"tt_img_upl_{seq}")

        if paste_result.image_data is not None:
            buf = io.BytesIO()
            paste_result.image_data.save(buf, format="PNG")
            st.session_state["tt_img_bytes"] = buf.getvalue()
            st.session_state["tt_img_mime"] = "image/png"
            st.session_state["tt_img_seq"] += 1  # 다음 붙여넣기가 되도록 컴포넌트 재마운트
        elif img:
            st.session_state["tt_img_bytes"] = img.getvalue()
            st.session_state["tt_img_mime"] = img.type

        image_bytes = st.session_state.get("tt_img_bytes")
        mime_type = st.session_state.get("tt_img_mime", "image/png")

        if image_bytes:
            st.image(image_bytes, caption="붙여넣은(또는 업로드한) 이미지", width=300)

        if image_bytes and st.button("이미지에서 강좌 인식하기", key="tt_img_go"):
            try:
                st.session_state["tt_img_results"] = parse_timetable_image(image_bytes, mime_type)
            except Exception as e:
                st.session_state["tt_img_results"] = []
                st.error(f"인식 실패: {e}")

        img_results = st.session_state.get("tt_img_results")
        if img_results:
            for i, c in enumerate(img_results):
                matched = match_real_session(c.get("subject", ""), c.get("start_date", ""))
                if matched:
                    weekend = _is_weekend_days(matched["days"])
                    label = (f"{matched['subject']} — {matched['start_time']} ({matched['room']}, {matched['teacher']}) "
                             f"개강 {matched['start_date']}" + ("  [주말]" if weekend else "") + "  ✅ 시간표에서 찾음")
                    nd, nt, ns = matched["start_date"], matched["start_time"], matched["subject"]
                else:
                    weekend = _is_weekend_days(_expand_days(c.get("day_label", "")))
                    label = (f"{c.get('subject','')} — {c.get('start_time','')} 개강 {c.get('start_date','')}"
                             + ("  [주말]" if weekend else "") + "  ⚠️ 시간표에서 못 찾음(인식값 그대로 사용)")
                    nd, nt, ns = c.get("start_date", ""), c.get("start_time", ""), c.get("subject", "")
                key = f"img_pick_{i}_{cb_seq}"
                st.checkbox(label, key=key)
                candidates.append(dict(key=key, nd=nd, nt=nt, ns=ns, nts="주말" if weekend else ""))
        elif img_results is not None:
            st.caption("인식된 강좌가 없어요.")

    if candidates:
        if st.button("✅ 체크한 강좌 배정 문구 생성 (여러 개 누적 가능)", use_container_width=True, key="gen_checked"):
            picked = [c for c in candidates if st.session_state.get(c["key"])]
            if picked:
                lines = [gen_text("신규", nd=c["nd"] or today.isoformat(), nt=c["nt"], ns=c["ns"], nts=c["nts"])
                          for c in picked]
                new_text = "\n".join(lines)
                ss.assign_out = (ss.assign_out + "\n" + new_text) if ss.assign_out else new_text
                st.session_state["cb_gen_seq"] += 1  # 체크박스 전부 재마운트해 초기화 (중복 누적 방지)
                st.rerun()
            else:
                st.caption("체크된 강좌가 없어요.")

    gtype = st.selectbox("배정 유형", ["신규 배정","과목변경 배정","배정 취소","날짜변경 배정"],
                         label_visibility="collapsed")
    result = ""

    if gtype == "신규 배정":
        with st.form("g1"):
            c1,c2,c3,c4 = st.columns(4)
            nd  = c1.date_input("날짜*", value=today, key="g1_nd")
            nt  = c2.text_input("시간*", placeholder="12:00", key="g1_nt")
            ns  = c3.text_input("과목명*", placeholder="스케치업2", key="g1_ns")
            nts = c4.text_input("시간대", placeholder="주말", key="g1_nts")
            if st.form_submit_button("✨ 문구 생성", use_container_width=True):
                result = gen_text("신규", nd=nd.isoformat(), nt=nt, ns=ns, nts=nts)

    elif gtype == "과목변경 배정":
        st.markdown("<div style='font-size:11px;color:var(--color-muted);margin-bottom:4px'>▸ 이전 강좌</div>", unsafe_allow_html=True)
        old_s = _course_picker("이전", "cg_old")
        st.markdown("<div style='font-size:11px;color:var(--color-muted);margin:8px 0 4px'>▸ 변경 후 강좌</div>", unsafe_allow_html=True)
        new_s = _course_picker("변경 후", "cg_new")
        if old_s and new_s:
            auto_ofee = lookup_fee(old_s["subject"], _is_weekend_days(old_s["days"])) or 0
            auto_nfee = lookup_fee(new_s["subject"], _is_weekend_days(new_s["days"])) or 0
            c1, c2 = st.columns(2)
            ofee = money_input(f"이전 수강료 (자동조회: {auto_ofee:,}원)", auto_ofee, "cg_ofee", c1)
            nfee = money_input(f"변경 후 수강료 (자동조회: {auto_nfee:,}원)", auto_nfee, "cg_nfee", c2)
            if st.button("✨ 문구 생성", use_container_width=True, key="cg_gen"):
                result = gen_text("과목변경", os_=old_s["subject"], ofee=int(ofee),
                                  nd=new_s["start_date"], nt=new_s["start_time"], ns=new_s["subject"],
                                  nts="주말" if _is_weekend_days(new_s["days"]) else "", nfee=int(nfee))

    elif gtype == "배정 취소":
        cancel_s = _course_picker("취소할", "cx")
        if cancel_s and st.button("✨ 문구 생성", use_container_width=True, key="cx_gen"):
            result = gen_text("취소", od=cancel_s["start_date"], ot=cancel_s["start_time"], os_=cancel_s["subject"])

    elif gtype == "날짜변경 배정":
        st.markdown("<div style='font-size:11px;color:var(--color-muted);margin-bottom:4px'>▸ 이전 일정</div>", unsafe_allow_html=True)
        old_s = _course_picker("이전", "dc_old")
        st.markdown("<div style='font-size:11px;color:var(--color-muted);margin:8px 0 4px'>▸ 변경 후 일정</div>", unsafe_allow_html=True)
        new_s = _course_picker("변경 후", "dc_new")
        if old_s and new_s and st.button("✨ 문구 생성", use_container_width=True, key="dc_gen"):
            result = gen_text("날짜변경", od=old_s["start_date"], ot=old_s["start_time"], os_=old_s["subject"],
                              nd=new_s["start_date"], nt=new_s["start_time"])

    if result: ss.assign_out = result
    if ss.assign_out:
        lines_n = ss.assign_out.count("\n") + 1
        st.text_area("📋 생성된 배정 문구 (복사하세요)", value=ss.assign_out, height=max(80, min(400, 30 * lines_n)))
        if st.button("🗑 초기화"):
            ss.assign_out = ""; st.rerun()


# ── 좌측 사이드바 nav (클로드놀이_배포판 index.html 앱셸) ──────────
with st.sidebar:
    st.markdown('<div class="sb-brand">■ SBS아카데미 대전</div>'
                '<div class="sb-brand-sub">업무 대시보드</div>', unsafe_allow_html=True)
    PAGES = ["🏠 홈", "✅ 체크리스트", "📅 강의시간표", "🎯 강의배정", "📄 개강안내문", "🗓️ 상담시간표", "📝 메모 양식", "🍚 점심값 정산"]
    page = st.radio("메뉴", PAGES, label_visibility="collapsed")

    # 녹취 메모 앱은 내부망 콜 시스템에 붙어야 해서 Cloud 통합 불가 → LAN 주소로 바로가기만
    st.divider()
    _nokchwi_url = get_state("nokchwi_url", "http://192.168.123.104:5000")
    st.link_button("🎙️ 녹취 메모 열기", _nokchwi_url, use_container_width=True)
    with st.expander("녹취 주소 변경"):
        _nu = st.text_input("녹취 앱 주소 (같은 네트워크)", _nokchwi_url, key="nokchwi_url_in",
                            help="녹취 앱(web_app.py)을 돌리는 PC 주소. 예 http://192.168.0.12:5000")
        if _nu.strip() and _nu.strip() != _nokchwi_url and st.button("저장", key="nokchwi_url_save"):
            set_state("nokchwi_url", _nu.strip())
            st.rerun()

# ── ✅ 체크리스트 (홈에서 분리 · fragment 인라인 체크) ────────────────
if page == "✅ 체크리스트":
    st.markdown("#### ✅ 체크리스트")
    st.caption("오늘 할 일 · 매일 한국시간 06:00에 완료 체크가 자동으로 풀립니다.")

    CL_CATS = [("우선순위1", "🔴 우선순위 1"), ("우선순위2", "🟠 우선순위 2"), ("주중업무", "🔵 주중업무")]
    _members = get_team_members() + ["미지정"]

    fc = st.columns([1.4, 1, 1])
    who = fc[0].selectbox("담당자", ["전체"] + _members, label_visibility="collapsed")
    edit_mode = fc[2].toggle("✏️ 편집", key="cl_edit")

    all_tasks = get_tasks()
    view = [t for t in all_tasks if who == "전체" or (t["assignee"] or "미지정") == who]
    done_n = sum(1 for t in view if t["is_done"])
    tot_n = len(view)
    pctv = round(done_n / tot_n * 100) if tot_n else 0
    fc[1].markdown(f"<div style='text-align:right;font-size:12px;color:var(--color-text-secondary);padding-top:6px'>"
                   f"{done_n} / {tot_n} 완료</div>", unsafe_allow_html=True)
    st.markdown(f"<div style='background:#EDF2F6;border-radius:6px;height:8px;overflow:hidden;margin:2px 0 14px'>"
                f"<div style='background:var(--color-primary);height:100%;width:{pctv}%'></div></div>",
                unsafe_allow_html=True)

    if edit_mode:
        # 편집 모드 — 표에서 제목·담당자·분류·날짜 일괄 수정
        import pandas as _pd
        rows = [{"id": t["id"], "완료": bool(t["is_done"]), "할 일": t["title"],
                 "담당자": t["assignee"] or "", "분류": t["category"],
                 "날짜": t["task_date"] or ""} for t in view]
        df = _pd.DataFrame(rows, columns=["id", "완료", "할 일", "담당자", "분류", "날짜"])
        ed = st.data_editor(
            df, key="cl_editor", hide_index=True, use_container_width=True, num_rows="fixed",
            column_config={
                "id": None,
                "완료": st.column_config.CheckboxColumn("✓"),
                "할 일": st.column_config.TextColumn("할 일", width="large"),
                "담당자": st.column_config.TextColumn("담당자", width="small"),
                "분류": st.column_config.SelectboxColumn("분류", options=[c[0] for c in CL_CATS]),
                "날짜": st.column_config.TextColumn("날짜(YYYY-MM-DD)", width="small"),
            },
        )
        if st.button("💾 편집 저장", type="primary"):
            orig = {r["id"]: r for r in rows}
            for _, r in ed.iterrows():
                o = orig.get(r["id"])
                if not o:
                    continue
                if (r["할 일"] != o["할 일"] or r["담당자"] != o["담당자"]
                        or r["분류"] != o["분류"] or r["날짜"] != o["날짜"]):
                    update_task(int(r["id"]), str(r["할 일"]).strip(), r["분류"],
                                str(r["날짜"]).strip(), str(r["담당자"]).strip())
                if bool(r["완료"]) != o["완료"]:
                    toggle_task(int(r["id"]), bool(r["완료"]))
            st.toast("저장됨"); st.rerun()
        with st.expander("👥 담당자 추가/삭제"):
            m1, m2 = st.columns([3, 1])
            nm = m1.text_input("이름", key="cl_new_m", label_visibility="collapsed", placeholder="담당자 이름")
            if m2.button("추가", key="cl_add_m") and nm.strip():
                add_team_member(nm.strip()); st.rerun()
            for m in get_team_members():
                r1, r2 = st.columns([3, 1])
                r1.caption(m)
                if r2.button("삭제", key=f"cl_del_m_{m}"):
                    remove_team_member(m); st.rerun()
        st.stop()

    @st.fragment
    def _checklist_body():
        cur = [t for t in get_tasks() if who == "전체" or (t["assignee"] or "미지정") == who]
        for cat_key, cat_label in CL_CATS:
            ct = [t for t in cur if t["category"] == cat_key]
            st.markdown(f"<div style='font-size:12px;font-weight:700;margin:14px 0 6px;color:var(--color-text-secondary)'>"
                        f"{cat_label} · {sum(1 for t in ct if t['is_done'])}/{len(ct)}</div>", unsafe_allow_html=True)

            if cat_key == "주중업무" and who != "전체" and who != "미지정":
                wp = get_weekday_pct(who, week_start)
                nwp = st.slider("주중 진행률", 0, 100, wp, step=5, key=f"cl_wp_{who}")
                if nwp != wp:
                    set_weekday_pct(who, week_start, nwp); st.rerun(scope="fragment")

            if not ct:
                st.markdown("<div style='color:var(--color-muted);font-size:12px;padding:4px 0 2px'>· 항목 없음</div>",
                            unsafe_allow_html=True)
            for t in ct:
                cc = st.columns([0.08, 0.84, 0.08])
                dv = cc[0].checkbox("완", value=bool(t["is_done"]), key=f"cl_ck_{t['id']}",
                                    label_visibility="collapsed")
                if dv != bool(t["is_done"]):
                    toggle_task(t["id"], dv); st.rerun(scope="fragment")
                dd = f" · {t['task_date'][5:].replace('-', '.')}" if t["task_date"] else ""
                nm = f" · {t['assignee']}" if (who == "전체" and t["assignee"]) else ""
                style = ("color:var(--color-muted);text-decoration:line-through" if t["is_done"]
                         else "color:var(--color-text)")
                cc[1].markdown(f"<div style='font-size:13.5px;padding-top:4px;{style}'>{t['title']}"
                               f"<span style='font-size:11px;color:var(--color-muted)'>{dd}{nm}</span></div>",
                               unsafe_allow_html=True)
                if cc[2].button("✕", key=f"cl_x_{t['id']}"):
                    del_task(t["id"]); st.rerun(scope="fragment")

        # 빠른 추가
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        with st.form("cl_add", clear_on_submit=True):
            a = st.columns([2.2, 1, 0.8])
            nt = a[0].text_input("할 일", placeholder="+ 새 할 일", label_visibility="collapsed", key="cl_nt")
            ncat = a[1].selectbox("분류", [c[0] for c in CL_CATS], label_visibility="collapsed", key="cl_ncat")
            if a[2].form_submit_button("추가", use_container_width=True) and nt.strip():
                asg = "" if who in ("전체", "미지정") else who
                add_task(nt.strip(), ncat, today.isoformat(), asg)
                st.rerun(scope="fragment")

    _checklist_body()
    st.stop()

# ── 점심값 정산 (가계부 스타일) ─────────────────────────────────
if page == "🍚 점심값 정산":
    st.markdown("#### 🍚 점심값 정산 장부")
    people = lunch_members()
    DOW = ["월", "화", "수", "목", "금", "토", "일"]

    # ── 주 이동 ──
    if "lunch_wk" not in ss:
        ss.lunch_wk = (today - timedelta(days=today.weekday())).isoformat()
    mon = date.fromisoformat(ss.lunch_wk)
    sun = mon + timedelta(days=6)
    nav = st.columns([1, 1, 2, 1])
    if nav[0].button("◀", key="lw_prev", use_container_width=True):
        ss.lunch_wk = (mon - timedelta(days=7)).isoformat(); st.rerun()
    if nav[1].button("오늘", key="lw_this", use_container_width=True):
        ss.lunch_wk = (today - timedelta(days=today.weekday())).isoformat(); st.rerun()
    nav[2].markdown(f"<div style='text-align:center;font-weight:700;font-size:15px;padding-top:5px'>"
                    f"{mon.year}. {mon.month:02d}. {mon.day:02d} ~ {sun.month:02d}. {sun.day:02d}</div>", unsafe_allow_html=True)
    if nav[3].button("▶", key="lw_next", use_container_width=True):
        ss.lunch_wk = (mon + timedelta(days=7)).isoformat(); st.rerun()

    days, data = get_lunch_week(ss.lunch_wk, people)
    day_total = get_lunch_days(ss.lunch_wk)                 # {ymd: total}
    paid_map = get_lunch_paid(ss.lunch_wk, people)          # {(person, ymd): bool}
    wk_paid = get_lunch_week_paid(ss.lunch_wk, people)      # {person: bool}
    date_lbl = [f"{date.fromisoformat(dy).month}/{date.fromisoformat(dy).day} ({dn})"
                for dy, dn in zip(days, DOW)]

    # ── 장부: 사람별 금액 + 그날 실제 결제액(직접 입력) ──
    ledger = pd.DataFrame({"날짜": date_lbl})
    for p in people:
        ledger[p] = [int(data[p][i]) for i in range(7)]
    ledger["그날 총액"] = [int(day_total[dy]) for dy in days]
    edited = st.data_editor(
        ledger, key="lunch_editor", hide_index=True, use_container_width=True, num_rows="fixed",
        disabled=["날짜"],
        column_config={
            "날짜": st.column_config.TextColumn("날짜", width="small", disabled=True),
            **{p: st.column_config.NumberColumn(f"{p} (원)", min_value=0, step=500, format="%d") for p in people},
            "그날 총액": st.column_config.NumberColumn("그날 총액(원)", min_value=0, step=1000, format="%d",
                                                  help="그날 식당에서 실제 결제한 금액 — 사람별 합계와 별개로 직접 입력"),
        },
    )
    cur = {p: [int(edited[p].iloc[i]) if pd.notna(edited[p].iloc[i]) else 0 for i in range(7)] for p in people}
    cur_total = [int(edited["그날 총액"].iloc[i]) if pd.notna(edited["그날 총액"].iloc[i]) else 0 for i in range(7)]

    # ── 사람별 냈음: 그날 자기 몫을 냈는지 (사람 × 날짜) ──
    st.caption("✅ 사람별 냈음 — 그날 자기 몫을 냈는지 체크")
    pledger = pd.DataFrame({"날짜": date_lbl})
    for p in people:
        pledger[p] = [bool(paid_map[(p, days[i])]) for i in range(7)]
    pedited = st.data_editor(
        pledger, key="lunch_paid_editor", hide_index=True, use_container_width=True, num_rows="fixed",
        disabled=["날짜"],
        column_config={
            "날짜": st.column_config.TextColumn("날짜", width="small", disabled=True),
            **{p: st.column_config.CheckboxColumn(f"{p}") for p in people},
        },
    )
    cur_pp = {(p, days[i]): bool(pedited[p].iloc[i]) for p in people for i in range(7)}

    # ── 사람별 주차 정산완료: 이번 주 정산액을 입금했는지 (사람 × 주차) ──
    st.caption("💸 주차 정산완료 — 이번 주 정산액을 입금했는지 체크")
    wck = st.columns(len(people))
    cur_wk = {p: wck[i].checkbox(p, value=wk_paid[p], key=f"lwk_{ss.lunch_wk}_{p}")
              for i, p in enumerate(people)}

    dirty = (any(cur[p][i] != data[p][i] for p in people for i in range(7))
             or any(cur_total[i] != day_total[days[i]] for i in range(7))
             or any(cur_pp[(p, dy)] != paid_map[(p, dy)] for p in people for dy in days)
             or any(cur_wk[p] != wk_paid[p] for p in people))

    sc = st.columns([1, 1, 3])
    if sc[0].button("💾 저장", type="primary", use_container_width=True, disabled=not dirty):
        for p in people:
            for dy, v in zip(days, cur[p]):
                save_lunch(p, dy, v)
        for i, dy in enumerate(days):
            save_lunch_day(dy, cur_total[i])
        for (p, dy), v in cur_pp.items():
            if v != paid_map[(p, dy)]:
                save_lunch_paid(p, dy, v)
        for p in people:
            if cur_wk[p] != wk_paid[p]:
                save_lunch_week_paid(p, ss.lunch_wk, cur_wk[p])
        st.toast("저장됐어요"); st.rerun()
    if sc[1].button("↩ 되돌리기", use_container_width=True, disabled=not dirty):
        for _k in ("lunch_editor", "lunch_paid_editor"):
            ss.pop(_k, None)
        for p in people:
            ss.pop(f"lwk_{ss.lunch_wk}_{p}", None)
        st.rerun()
    if dirty:
        sc[2].markdown("<div style='color:var(--color-urgent);font-size:12px;padding-top:8px'>✏️ 수정됨 — 저장해야 남습니다</div>", unsafe_allow_html=True)

    # ── 합계 (장부 하단 줄) ──
    totals = {p: sum(cur[p]) for p in people}
    daily = [sum(cur[p][i] for p in people) for i in range(7)]
    grand = sum(totals.values())
    n = max(len(people), 1)
    share = grand // n
    paid_sum = sum(cur_total)                                   # 직접 입력한 그날 총액의 주간 합
    owe_day = [{p for p in people if cur[p][i] > 0} for i in range(7)]
    got_day = [{p for p in people if cur_pp[(p, days[i])]} for i in range(7)]

    def _day_status(i):
        o, g = owe_day[i], got_day[i]
        if o and o <= g:
            return "full"
        return "partial" if g else None

    split_mode = st.radio("정산 방식", ["n빵 (총액 ÷ 인원)", "각자 자기 것만"], horizontal=True, key="lunch_mode")
    owe_amt = {p: (share if split_mode.startswith("n빵") else totals[p]) for p in people}
    unpaid_people = [p for p in people if not cur_wk[p] and owe_amt[p] > 0]
    unpaid_amt = sum(owe_amt[p] for p in unpaid_people)

    foot = "<table style='width:100%;border-collapse:collapse;font-size:13px;margin-top:6px'>"
    foot += "<tr style='background:var(--color-selected)'><th style='text-align:left;padding:7px 10px'>사람별 합계</th>"
    for i in range(7):
        foot += f"<td style='text-align:right;padding:7px 8px;color:var(--color-text-secondary)'>{daily[i]:,}</td>"
    foot += "</tr>"
    foot += "<tr><th style='text-align:left;padding:7px 10px'>그날 결제액</th>"
    for i in range(7):
        stt = _day_status(i)
        mark = ("<span style='color:var(--color-success)'> ✓</span>" if stt == "full"
                else "<span style='color:var(--color-muted)'> 부분</span>" if stt == "partial" else "")
        foot += f"<td style='text-align:right;padding:7px 8px;color:var(--color-text-secondary)'>{cur_total[i]:,}{mark}</td>"
    foot += "</tr></table>"
    st.markdown(foot, unsafe_allow_html=True)

    # ── 이번 주 정산 카드 ──
    st.markdown("##### 이번 주 정산")
    cards = st.columns(len(people) + 1)
    for i, p in enumerate(people):
        cards[i].metric(f"{p} 낸 돈", f"{totals[p]:,}원")
    cards[-1].metric("총 점심값", f"{grand:,}원", help=f"÷{n} = 1인당 {share:,}원")

    mc = st.columns(4)
    mc[0].metric("이번 주 실제 결제 합계", f"{paid_sum:,}원", help="장부에 직접 입력한 '그날 총액'의 주간 합")
    mc[1].metric("이번 주 미정산 인원", f"{len(unpaid_people)}명",
                 help=(", ".join(unpaid_people) if unpaid_people else "전원 정산완료"))
    mc[2].metric("미납 금액 합계", f"{unpaid_amt:,}원", help="주차 정산 미완료자들의 정산액 합")
    if paid_sum and paid_sum != grand:
        mc[3].metric("사람별 합계와 차이", f"{paid_sum - grand:+,}원",
                     help=f"사람별 합계 {grand:,}원 · 실제 결제 {paid_sum:,}원")

    if split_mode.startswith("n빵"):
        st.caption(f"총 {grand:,}원 ÷ {n}명 = 1인당 **{share:,}원**")
        rows = ""
        for p in people:
            bal = totals[p] - share
            if bal > 0:
                tag = f"<b style='color:var(--color-success)'>+{bal:,}원 받을 돈</b>"
            elif bal < 0:
                tag = f"<b style='color:var(--color-urgent)'>{-bal:,}원 더 내야 함</b>"
            else:
                tag = "<span style='color:var(--color-muted)'>정산 완료</span>"
            wtag = ("<b style='color:var(--color-success)'>✅ 정산완료</b>" if cur_wk[p]
                    else "<span style='color:var(--color-urgent)'>⬜ 미정산</span>")
            rows += (f"<tr><td style='padding:8px 10px;border-bottom:1px solid var(--color-border)'>{p}</td>"
                     f"<td style='padding:8px 10px;text-align:right;border-bottom:1px solid var(--color-border);color:var(--color-text-secondary)'>낸 돈 {totals[p]:,} · 몫 {share:,}</td>"
                     f"<td style='padding:8px 10px;text-align:right;border-bottom:1px solid var(--color-border)'>{tag}</td>"
                     f"<td style='padding:8px 10px;text-align:right;border-bottom:1px solid var(--color-border)'>{wtag}</td></tr>")
        st.markdown(f"<table style='width:100%;border-collapse:collapse;font-size:13px;background:var(--color-surface);"
                    f"border:1px solid var(--color-border);border-radius:10px;overflow:hidden'>{rows}</table>",
                    unsafe_allow_html=True)
        recv = [p for p in people if totals[p] - share > 0]
        if recv:
            st.markdown(f"→ **{', '.join(recv)}** 님에게 입금하면 정산 끝")
    else:
        st.caption("각자 자기가 낸 만큼만 입금")
        rows = "".join(f"<tr><td style='padding:8px 10px;border-bottom:1px solid var(--color-border)'>{p}</td>"
                       f"<td style='padding:8px 10px;text-align:right;border-bottom:1px solid var(--color-border)'><b>{totals[p]:,}원</b></td>"
                       f"<td style='padding:8px 10px;text-align:right;border-bottom:1px solid var(--color-border)'>"
                       + ("<b style='color:var(--color-success)'>✅ 정산완료</b>" if cur_wk[p]
                          else "<span style='color:var(--color-urgent)'>⬜ 미정산</span>")
                       + "</td></tr>"
                       for p in people)
        st.markdown(f"<table style='width:100%;border-collapse:collapse;font-size:13px;background:var(--color-surface);"
                    f"border:1px solid var(--color-border);border-radius:10px;overflow:hidden'>{rows}</table>",
                    unsafe_allow_html=True)

    # ── 복사용 (카톡 붙여넣기) ──
    with st.expander("📋 카톡에 붙여넣을 정산 문구"):
        lines = [f"[점심값 정산] {mon.month}/{mon.day}~{sun.month}/{sun.day}", ""]
        for p in people:
            lines += [p, " ".join(DOW), "/".join(f"{v:,}원" for v in cur[p]), f"주 합계: {totals[p]:,}원", ""]
        lines.append("─" * 18)
        lines.append(f"총액: {' + '.join(f'{totals[p]:,}' for p in people)} = {grand:,}원")
        if split_mode.startswith("n빵"):
            lines.append(f"1인당(÷{n}): {share:,}원\n")
            for p in people:
                bal = totals[p] - share
                lines.append(f"{p}: {totals[p]:,} 냄 → " +
                             (f"+{bal:,}원 받음" if bal > 0 else (f"-{-bal:,}원 냄" if bal < 0 else "정산완료")))
        else:
            lines.append("각자 입금: " + " / ".join(f"{p} {totals[p]:,}원" for p in people))
        if paid_sum:
            diff = paid_sum - grand
            lines.append("")
            lines.append(f"실제 결제 총액: {paid_sum:,}원" + ("" if diff == 0 else f" (사람별 합계와 {diff:+,}원)"))
        lines.append("")
        lines.append("정산 현황 — " + " / ".join(
            f"{p} " + ("✅" if cur_wk[p] else f"⬜({owe_amt[p]:,}원)") for p in people))
        st.code("\n".join(lines), language=None)

    # ── 월 달력 ──
    with st.expander(f"📅 {mon.year}년 {mon.month}월 — 날짜별 점심값"):
        _ym = f"{mon.year:04d}-{mon.month:02d}"
        mt = lunch_month_totals(_ym)            # {day: 사람별 amount 합}
        mdt = lunch_month_days(_ym)             # {day: 그날 총액 입력값}
        mps = lunch_month_paid(_ym, people)     # {day: 'full'|'partial'}
        head = "".join(f"<th style='padding:5px;font-size:11px;color:var(--color-muted);font-weight:500'>{d}</th>" for d in DOW)
        body = ""
        for wk in cal_lib.monthcalendar(mon.year, mon.month):
            body += "<tr>"
            for dd in wk:
                if dd == 0:
                    body += "<td style='border:1px solid var(--color-border);height:58px;background:var(--color-bg)'></td>"
                else:
                    amt = mdt.get(dd) or mt.get(dd, 0)
                    stt = mps.get(dd)
                    if stt == "full":
                        col, chk = "var(--color-success)", "<span style='color:var(--color-success)'> ✓</span>"
                    elif stt == "partial":
                        col, chk = "var(--color-muted)", "<span style='color:var(--color-muted)'> ◐</span>"
                    else:
                        col, chk = "var(--color-primary)", ""
                    inr = (f"<span style='font-size:10px;color:var(--color-muted)'>{dd}</span>"
                           + (f"<div style='color:{col};font-weight:700;font-size:12px;margin-top:4px'>{amt:,}{chk}</div>" if amt else ""))
                    body += f"<td style='border:1px solid var(--color-border);height:58px;padding:5px;vertical-align:top'>{inr}</td>"
            body += "</tr>"
        st.markdown(f"<table style='border-collapse:collapse;width:100%;table-layout:fixed'><tr>{head}</tr>{body}</table>",
                    unsafe_allow_html=True)

    # ── 설정 ──
    with st.expander("⚙️ 인원 설정"):
        _nm = st.text_input("이름 (쉼표로 구분)", ", ".join(people), key="lunch_nm")
        if st.button("저장", key="lunch_nm_save"):
            set_lunch_members([x.strip() for x in _nm.split(",") if x.strip()]); st.rerun()
    st.stop()

# ── 📝 메모 양식 (카테고리별 텍스트 양식 · 담당자 자동 치환 · 팀 공유 편집) ──
if page == "📝 메모 양식":
    st.markdown("#### 📝 메모 양식")
    st.caption("컨택메모·상담완료·예약·안내메세지 등 12종. 양식은 팀이 함께 편집·저장하고, "
               "담당자 이름·전화는 **볼 때·복사할 때 자동으로** 채워집니다. 저장은 플레이스홀더가 남은 원본으로 됩니다.")

    _cs = get_consultants()
    _mc1, _mc2 = st.columns([1, 1.4])
    _msel = _mc1.selectbox("담당자", list(range(len(_cs))),
                           format_func=lambda i: (_cs[i]["name"] or "(이름 미입력)"), key="memo_consultant_sel")
    _mconsult = dict(_cs[_msel])
    if not (_mconsult.get("name") or "").strip():
        _mc1.caption("담당자 이름이 비어 있어요 → 🗓️ 상담시간표의 ‘담당자 명단 관리’에서 채우면 여기에도 반영됩니다. "
                     "지금은 플레이스홀더 그대로 나옵니다.")
    _mkey = _mc2.selectbox("양식 종류", [k for k, _ in MEMO_LABELS],
                           format_func=lambda k: dict(MEMO_LABELS)[k], key="memo_cat_sel")

    _ta_key = f"memo_ta_{_mkey}"
    _raw = st.text_area("양식 본문 (편집·저장 대상 · 플레이스홀더 유지)", value=get_memo_tpl(_mkey),
                        height=430, key=_ta_key)

    _b1, _b2, _b3 = st.columns([1, 1, 4])
    if _b1.button("💾 양식 저장", key="memo_save", use_container_width=True):
        save_memo_tpl(_mkey, _raw)
        ss.pop(_ta_key, None)
        st.success("양식 저장됨 (팀 공유)"); st.rerun()
    if _b2.button("↩️ 기본값으로", key="memo_reset", use_container_width=True):
        reset_memo_tpl(_mkey)
        ss.pop(_ta_key, None)
        st.success("기본 양식으로 되돌렸습니다"); st.rerun()

    st.markdown("**📋 복사용 (담당자·전화 반영됨)** — 오른쪽 위 복사 아이콘")
    st.code(memo_fill(_raw, _mconsult), language=None)
    st.stop()

# 시간표 4개 메뉴 → 개강안내.html 을 각각 embed(해시로 탭 지정). 각 iframe이 별도라
# 상태는 안 섞이므로, DB에 저장한 업로드분(tt_upload)을 preload로 4곳 모두에 주입해 같은 시간표를 쓴다.
_TT_MENU = {"📅 강의시간표": "timetable", "🎯 강의배정": "assign",
            "📄 개강안내문": "announce", "🗓️ 상담시간표": "consult"}
if page in _TT_MENU:
    _hash = _TT_MENU[page]
    _ups = tt_uploads()
    _sample = os.path.join(os.path.dirname(__file__), "시간표도구", "data", "시간표.xls")
    if not _ups and os.path.exists(_sample) and st.checkbox("샘플 시간표로 먼저 써보기 (data/시간표.xls)", key="tt_sample"):
        import base64 as _sb
        with open(_sample, "rb") as _sf:
            _ups = [("샘플_시간표.xls", _sb.b64encode(_sf.read()).decode())]

    with st.expander(f"📤 강의시간표 엑셀 업로드 · 팀 공유  ({len(_ups)}개)", expanded=not _ups):
        st.caption("여기 올린 시간표 하나가 **4개 메뉴 전부 + 홈의 배정/문자 기능**에 함께 쓰입니다. "
                   "인트라넷에서 받은 `[AIX대전]강의시간표*.xls` / `[컴퓨터대전]*.xls`를 올리세요. DB 저장 → 팀원 모두 반영.")
        up = st.file_uploader("강의시간표 xls", type=["xls", "xlsx"], accept_multiple_files=True,
                              key="tt_up", label_visibility="collapsed")
        if up:
            for f in up:
                save_tt_upload(f.name, f.getvalue())
            st.success(f"{len(up)}개 저장됨"); st.rerun()
        for name, _b64 in _ups:
            cx, cy = st.columns([1, 0.12])
            cx.markdown(f"<div style='font-size:12px;padding:4px 0'>· {name}</div>", unsafe_allow_html=True)
            if name != "샘플_시간표.xls" and cy.button("✕", key=f"delu_{name}"):
                del_tt_upload(name); st.rerun()
        if _ups and any(n != "샘플_시간표.xls" for n, _ in _ups):
            if st.button("전체 삭제 (인트라넷 시간표/ 폴더 기본값으로)"):
                clear_tt_uploads(); st.rerun()
        elif not _ups:
            st.info(f"업로드 없음 — 저장소 `인트라넷 시간표/` 폴더의 파일 {len(_tt_folder_b64())}개를 자동으로 씁니다 (팀원 모두 동일). "
                    "따로 올리면 그게 우선 적용됩니다.")

    # 임베드 도구에 넣을 파일: DB 업로드분 우선, 없으면 인트라넷 시간표/ 폴더 파일 (커밋돼 팀 공유)
    _preload = _ups or _tt_folder_b64()

    # ── 담당자(문서 서명) 선택 + 명단 관리 · 팀 공유 ── (상담시간표 / 개강안내문 / 견적서에 주입)
    _consultant = None
    if _hash in ("consult", "announce"):
        _cs = get_consultants()
        _sel = st.selectbox("담당자", list(range(len(_cs))),
                            format_func=lambda i: (_cs[i]["name"] or "(이름 미입력)"), key="tt_consultant_sel")
        _cd = st.date_input("기준일 (상담일 / 견적서 상담 일자)", value=date.today(), key="tt_consultant_date")
        _consultant = dict(_cs[_sel])
        _consultant["date"] = _cd.isoformat() if _cd else ""
        with st.expander("👤 담당자 명단 관리 (팀 공유 · DB 저장)"):
            st.caption("이름은 직함까지 넣으세요 (예: `홍길동 교육멘토`). 저장한 담당자를 위에서 골라 "
                       "상담시간표·개강안내문·견적서 서명에 넣습니다. 도구를 직접 열면 기존 기본값이 쓰입니다.")
            _cdf = pd.DataFrame(_cs, columns=["name", "phone", "mobile", "email"])
            _ced = st.data_editor(_cdf, num_rows="dynamic", use_container_width=True, hide_index=True,
                                  key="consultant_editor",
                                  column_config={"name": st.column_config.TextColumn("이름(직함 포함)", width="medium"),
                                                 "phone": "유선전화", "mobile": "휴대폰", "email": "이메일"})
            if st.button("💾 명단 저장", key="consultant_save"):
                save_consultants(_ced.fillna("").to_dict("records"))
                st.success("명단 저장됨 (팀 공유)"); st.rerun()

    _tips = {
        "timetable": "월/평일·주말 필터로 강의를 봅니다. 강의를 선택해 문서를 만들려면 **강의배정 / 개강안내문 / 상담시간표** 메뉴로 가세요.",
        "assign": "강의를 여러 개 골라 → 하단 막대 **배정 진행**.",
        "announce": "강의 하나 클릭 → 개강안내문(복사). 여러 개 골라 하단 막대 **개강안내문 생성**도 가능.",
        "consult": "강의를 **여러 칸 토글 선택** → 화면 하단 막대 **시간표 생성** → 미리보기에서 **인쇄 / PNG 저장**.",
    }
    st.caption(_tips[_hash] + " 인쇄·PNG가 이 화면에서 막히면 아래 **‘전체 화면 새 탭’** 링크로 여세요.")

    # 전체 화면 새 탭 링크 (같은 도구 + 시간표 파일 주입) — iframe sandbox에서 인쇄/PNG 막힐 때
    import base64 as _b64m
    _full = _tt_tool_html()
    if _consultant:
        _full += "<script>" + _consultant_inject(_consultant) + "</script>"
    if _preload:
        _fjs = json.dumps([{"name": n, "b64": b} for n, b in _preload], ensure_ascii=False)
        _full += ("<script>(function(){var FS=" + _fjs + ";async function p(){try{"
                  "if(typeof loadFileAuto!=='function'){setTimeout(p,150);return;}"
                  "for(var i=0;i<FS.length;i++){var b=atob(FS[i].b64),a=new Uint8Array(b.length);"
                  "for(var k=0;k<b.length;k++)a[k]=b.charCodeAt(k);await loadFileAuto(new File([a],FS[i].name));}"
                  "if(typeof saveToStorage==='function')saveToStorage();"
                  "location.hash='#" + _hash + "';if(typeof applyHashTab==='function')applyHashTab();"
                  "}catch(e){console.warn(e);}}if(document.readyState==='complete')p();else addEventListener('load',function(){setTimeout(p,200);});})();</script>")
    _uri = "data:text/html;base64," + _b64m.b64encode(_full.encode("utf-8")).decode()
    st.markdown(f"<a href='{_uri}' target='_blank' rel='noopener' "
                "style='display:inline-block;margin:2px 0 8px;font-size:12px;font-weight:600;color:var(--color-primary);"
                "border:1px solid var(--color-border);border-radius:8px;padding:6px 12px;text-decoration:none'>"
                "🔗 전체 화면 새 탭에서 열기</a>", unsafe_allow_html=True)

    if _hash == "assign":
        # 홈에서 옮겨온 '시간표 배정 자동화' (이미지/개인시간표 인식 → 배정 문구 생성)
        render_assign_automation()
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        st.caption("↓ 클로드놀이_배포판 강의배정 화면 (칸 선택 → 하단 막대 '배정 진행')")

    if _preload and not _ups:
        st.caption(f"⏳ `인트라넷 시간표/` 폴더 {len(_preload)}개 파일을 도구에 불러오는 중 — 목록이 뜨기까지 몇 초 걸릴 수 있습니다.")
    render_timetable_tool(_hash, preload=_preload or None, consultant=_consultant)
    st.stop()

# ── 헤더 (실시간 시계, 한국시간 기준) ────────────────────────────
st.iframe("""
<html><body style="margin:0;background:#F5F7F9">
<div style="font-family:'Pretendard','Noto Sans KR',sans-serif;display:flex;justify-content:space-between;
            align-items:center;padding:4px 2px 14px 2px;border-bottom:2px solid #E7EBEF">
  <h1 style="font-size:22px;font-weight:600;color:#333D46;margin:0">
    <span style="color:#5C7A94">■</span> 업무 및 일정 관리 대시보드
  </h1>
  <div id="live-clock" style="font-size:13px;color:#707F8D"></div>
</div>
<script>
function updateClock() {
  var now = new Date();
  var dOpts = {timeZone:'Asia/Seoul', year:'numeric', month:'long', day:'numeric', weekday:'short'};
  var dateStr = now.toLocaleDateString('ko-KR', dOpts);
  var timeStr = now.toLocaleTimeString('ko-KR', {timeZone:'Asia/Seoul', hour12:false});
  document.getElementById('live-clock').innerHTML =
    dateStr + '&nbsp;<span style="font-size:22px;font-weight:200;font-family:monospace;color:#333">' + timeStr + '</span>';
}
updateClock();
setInterval(updateClock, 1000);
</script>
</body></html>
""", height=70)

# ── 홈: 캘린더는 전체 폭으로 크게(1차), 상담·투두는 그 아래(2차) ──────
left = st.container()
right = st.container()

# ── 달력 (전체 폭) ───────────────────────────────────────────────
with left:
    ci_cls = ["ci1","ci2","ci3"]
    day_clr = [None,None,None,None,None,"#5C7A94","#C68A7A"]

    def pill_html(c, cls):
        rev = f"{c['expected_revenue']//10000}만" if c["expected_revenue"] >= 10000 \
              else (f"{c['expected_revenue']}원" if c["expected_revenue"] else "")
        sub = " ".join(x for x in [c["sched_time"], rev] if x)
        tip = f"{c['name']} {c['sched_time']} {c['expected_revenue']:,}원 ({c['ctype']})"
        return f"""<span class="ci {cls}" title="{tip}">
<b>{c['name']}</b>{' · '+sub if sub else ''}</span>"""

    rows_html = ""
    for week in cal_lib.monthcalendar(yr, mo):
        rows_html += "<tr style='display:contents'>"
        for wi, day in enumerate(week):
            if day == 0:
                rows_html += '<div class="cal-day" style="background:#F1F4F6"></div>'; continue
            is_td   = day == today.day
            td_cls  = " td" if is_td else ""
            nc      = "td" if is_td else ""
            dc      = day_clr[wi]
            nstyle  = f"color:{dc}" if dc and not is_td else ""
            events  = cmap.get(day, [])
            evhtml  = "".join(
                pill_html(e, ci_cls[i%3])
                for i,e in enumerate(events[:3])
            )
            rows_html += f"""
<div class="cal-day{td_cls}">
  <span class="cal-num {nc}" style="{nstyle}">{day}</span>
  {evhtml}
</div>"""

    st.markdown(f"""
<div class="db-card">
  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:14px'>
    <div class="db-card-title" style="margin:0">🗓️ 캘린더 일정</div>
    <div style='font-size:13px;font-weight:600;color:var(--color-text-secondary)'>{yr}년 {mo}월</div>
  </div>
  <div class="cal-grid">
    <div class="cal-head">월</div><div class="cal-head">화</div>
    <div class="cal-head">수</div><div class="cal-head">목</div>
    <div class="cal-head">금</div>
    <div class="cal-head" style="color:#5C7A94">토</div>
    <div class="cal-head" style="color:#C68A7A">일</div>
    {rows_html}
  </div>
</div>""", unsafe_allow_html=True)

    # 상담 추가 폼
    with st.expander("＋ 상담 일정 추가"):
        with st.form("cf", clear_on_submit=True):
            r1 = st.columns(2)
            cname = r1[0].text_input("이름 *")
            cdate = r1[1].date_input("날짜", value=today)
            r2 = st.columns([1, 1, 1])
            ctime  = r2[0].text_input("시간", placeholder="14:00")
            crev   = money_input("예정매출(원)", 0, "cf_crev", r2[1])
            ctype  = r2[2].selectbox("구분", ["단과", "정규"])
            cvisit = st.radio("방문 유형", ["신규방문", "재방문", "온라인"], horizontal=True)
            cassignee = st.text_input("담당자", placeholder="담당자")
            if st.form_submit_button("저장", use_container_width=True) and cname.strip():
                add_consult(cname.strip(), cdate.isoformat(), ctime, int(crev), ctype, cassignee.strip(), cvisit); st.rerun()

# ── RIGHT: 상담 일정(날짜별) + 칸반 TO DO LIST ──────────────────
with right:
    show_finalized = st.checkbox("완료 처리된 상담도 보기", key="show_finalized_consults")
    # 날짜 제한 없이(지난 상담 포함) 표시 — 지난 상담이 결과 체크 없이 넘어가면
    # 완료 처리를 못 해서 캘린더에 계속 남는 문제가 있었음
    display_consults = sorted(
        (c for c in consults if show_finalized or not c["finalized"]),
        key=lambda c: (c["sched_date"], c["sched_time"]))
    by_date = {}
    for c in display_consults:
        by_date.setdefault(c["sched_date"], []).append(c)
    # 마감보고 등 다른 곳에서 쓰는 "완료 숨김 반영된" 목록은 별도로 유지
    upcoming = sorted((c for c in consults if not c["finalized"]),
                       key=lambda c: (c["sched_date"], c["sched_time"]))

    st.markdown('<div class="db-card" style="margin-bottom:14px"><div class="db-card-title">🔵 상담 일정</div>',
                unsafe_allow_html=True)
    if not by_date:
        st.markdown('<div style="color:var(--color-muted);font-size:11px;text-align:center;padding:12px 0">예정된 상담 없음</div>',
                    unsafe_allow_html=True)
    for d in sorted(by_date):
        dd  = datetime.strptime(d, "%Y-%m-%d")
        wtag = ["월","화","수","목","금","토","일"][dd.weekday()]
        label = f"{dd.month}월 {dd.day}일 ({wtag}){' · 오늘' if d == today_str else ''}"
        clr = "#5C7A94" if d == today_str else "#999"
        st.markdown(f"<div style='font-size:11px;font-weight:700;color:{clr};margin:6px 0 4px'>{label}</div>",
                    unsafe_allow_html=True)
        for c in by_date[d]:
            if st.session_state.get("edit_consult_id") == c["id"]:
                with st.form(f"editcf_{c['id']}"):
                    e1, e2 = st.columns(2)
                    ename = e1.text_input("이름", value=c["name"])
                    edate = e2.date_input("날짜", value=datetime.strptime(c["sched_date"], "%Y-%m-%d").date())
                    e3, e4, e5 = st.columns(3)
                    etime = e3.text_input("시간", value=c["sched_time"])
                    erev  = money_input("예정매출(원)", c["expected_revenue"], f"erev_{c['id']}", e4)
                    ectype = e5.selectbox("구분", ["단과", "정규"], index=0 if c["ctype"] == "단과" else 1)
                    visit_opts = ["신규방문", "재방문", "온라인"]
                    evisit = st.radio("방문 유형", visit_opts, horizontal=True,
                                       index=visit_opts.index(c["visit_type"]) if c["visit_type"] in visit_opts else 0)
                    eassignee = st.text_input("담당자", value=c["assignee"])
                    s1, s2 = st.columns(2)
                    if s1.form_submit_button("저장", use_container_width=True, key=f"save_ec_{c['id']}"):
                        update_consult(c["id"], ename.strip(), edate.isoformat(), etime,
                                       int(erev), ectype, eassignee.strip(), evisit)
                        st.session_state["edit_consult_id"] = None
                        st.rerun()
                    if s2.form_submit_button("취소", use_container_width=True, key=f"cancel_ec_{c['id']}"):
                        st.session_state["edit_consult_id"] = None
                        st.rerun()
            else:
                cc1, cc2, cc3, cc4 = st.columns([1, 0.13, 0.13, 0.13])
                dim = "opacity:0.5;text-decoration:line-through" if c["finalized"] else ""
                cc1.markdown(f"""
<div style='background:#EDF2F6;border-left:3px solid #5C7A94;border-radius:6px;
            padding:7px 12px;margin:3px 0;display:flex;justify-content:space-between;align-items:center;{dim}'>
  <span style='font-size:12px;font-weight:700;color:#333D46'>{c['name']} <span style='font-weight:500;color:#707F8D'>({c['ctype']})</span></span>
  <span style='font-size:11px;color:#6D8AA3'>{c['sched_time']}</span>
  <span style='font-size:11px;font-weight:700;color:#5C7A94'>{c['expected_revenue']:,}원</span>
</div>""", unsafe_allow_html=True)
                if cc2.button("✏️", key=f"ec{c['id']}"):
                    st.session_state["edit_consult_id"] = c["id"]
                    st.rerun()
                if cc3.button("✕", key=f"dc{c['id']}"):
                    del_consult(c["id"]); st.rerun()
                if c["finalized"]:
                    if cc4.button("👁", key=f"unfin_{c['id']}", help="다시 보이게"):
                        unfinalize_consult(c["id"]); st.rerun()
                else:
                    if cc4.button("🙈", key=f"quickfin_{c['id']}", help="완료 처리(숨기기)"):
                        finalize_consult(c["id"]); st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    with st.expander("📋 상담 결과 입력 (등록/COD/미등록)", expanded=False):
        if not upcoming:
            st.caption("예정된 상담 없음")
        for c in upcoming:
            st.markdown(f"<div style='font-size:12px;font-weight:600;margin:6px 0 2px'>"
                        f"{c['sched_date'][5:].replace('-','.')} {c['sched_time']} {c['name']}</div>",
                        unsafe_allow_html=True)
            STATUS_OPTS = ["미정", "등록", "COD", "미등록"]
            cur_status = c["result_status"] if c["result_status"] in STATUS_OPTS else "미정"
            sc1, sc2 = st.columns([2, 1])
            new_status = sc1.radio("결과", STATUS_OPTS, index=STATUS_OPTS.index(cur_status),
                                    horizontal=True, key=f"cstat_{c['id']}", label_visibility="collapsed")
            if new_status != cur_status:
                set_consult_status(c["id"], new_status); st.rerun()
            if new_status in ("등록", "COD"):
                cur_amt = c["actual_amount"] or c["expected_revenue"]
                new_amt = money_input("실제 매출", cur_amt, f"camt_{c['id']}", sc2, label_visibility="collapsed")
                if new_amt != cur_amt:
                    set_consult_amount(c["id"], int(new_amt)); st.rerun()
            if new_status != "미정":
                if st.button("✅ 완료 (목록·캘린더에서 숨기기)", key=f"fin_{c['id']}", use_container_width=True):
                    finalize_consult(c["id"]); st.rerun()

            new_pct = st.slider("입금(등록) 예상 확률", 0, 100, c["deposit_pct"], step=5,
                                 key=f"cpct_{c['id']}")
            if new_pct != c["deposit_pct"]:
                set_consult_pct(c["id"], new_pct); st.rerun()


st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)



st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

# ── 개강 안내 문자 생성 ────────────────────────────────────────
st.markdown("""
<div class="db-card">
  <div class="db-card-title">📨 개강 안내 문자 생성</div>
  <div style='font-size:12px;color:var(--color-muted);margin-top:-10px;margin-bottom:14px'>
    강좌를 고르면 학생에게 보낼 개강 안내 문구가 자동으로 채워집니다.
  </div>
</div>""", unsafe_allow_html=True)

WEEKDAY_KO = "월화수목금토일"

def _hours_between(start, end):
    sh, sm = map(int, start.split(":"))
    eh, em = map(int, end.split(":"))
    diff = (eh * 60 + em) - (sh * 60 + sm)
    hours = diff / 60
    return int(hours) if hours == int(hours) else round(hours, 1)

notice_s = _course_picker("개강 안내할", "notice_pick")

if notice_s:
    room_floors = json.loads(get_state("room_floors", "{}"))
    floor_default = room_floors.get(notice_s["room"], "")
    nc1, nc2, nc3 = st.columns(3)
    sender_name = nc1.text_input("발신자(팀장) 이름", value=get_state("notice_sender_name", "송성기"), key="notice_sender_name_in")
    sender_phone = nc2.text_input("연락처", value=get_state("notice_sender_phone", "042-331-7511"), key="notice_sender_phone_in")
    branch_name = nc3.text_input("지점명", value=get_state("notice_branch_name", "SBS아카데미대전지점"), key="notice_branch_name_in")
    floor = st.text_input(f"'{notice_s['room']}' 강의장 층수", value=floor_default, placeholder="9층", key="notice_floor_in")

    if st.button("✨ 개강 안내 문구 생성", use_container_width=True, key="notice_gen"):
        set_state("notice_sender_name", sender_name.strip())
        set_state("notice_sender_phone", sender_phone.strip())
        set_state("notice_branch_name", branch_name.strip())
        if floor.strip():
            room_floors[notice_s["room"]] = floor.strip()
            set_state("room_floors", json.dumps(room_floors, ensure_ascii=False))

        sd = datetime.strptime(notice_s["start_date"], "%Y-%m-%d").date()
        ed = datetime.strptime(notice_s["end_date"], "%Y-%m-%d").date()
        hours = _hours_between(notice_s["start_time"], notice_s["end_time"])
        room_label = f"{notice_s['room']} 강의장" + (f" ({floor.strip()})" if floor.strip() else "")

        ss.notice_out = f"""안녕하세요, {branch_name} {sender_name} 팀장입니다.

아래 배정된 교육 과정 개강 일정이오니
스케줄 확인 해주시고, 확인 후 답변 부탁드립니다:)

■ 개강 일정

· 수강과목 : {notice_s['subject']}
· 개강일 : {sd.isoformat()} ({WEEKDAY_KO[sd.weekday()]})
· 종강일 : {ed.isoformat()} ({WEEKDAY_KO[ed.weekday()]})
· 요일 : {notice_s['day_label']} (주 {len(notice_s['days'])}일)
· 시간 : {notice_s['start_time']}~{notice_s['end_time']} ({hours}H)
· 강의장 : {room_label}


* 강의장은 개강 전날까지 변경될 수 있습니다.
* 코리아교육그룹-스마트러닝(앱/어플)에서 일정 및 강의장 확인이 가능합니다.
* 일정 변경 및 취소는 개강 3일 전까지만 가능하며, 수업 자리는 지정석이 아닙니다.
* 비대면 수업 링크는 문자 발송됩니다.

* 본 건물 지하주차장은 시간대에 따라 진출입 지연이 발생할 수 있으므로 가급적 대중교통 이용을 권장드립니다.
* 주차는 수업 시간만큼만 무료 지원되며, 초과 시 30분당 1,200원의 추가요금이 발생합니다. (추가요금은 실물카드 결제만 가능 / 삼성페이 불가)

☎{sender_phone}
{branch_name} {sender_name} 팀장"""
        st.rerun()

if ss.notice_out:
    lines_n = ss.notice_out.count("\n") + 1
    st.text_area("📋 생성된 개강 안내 문구 (복사하세요)", value=ss.notice_out, height=max(200, min(500, 22 * lines_n)))
    if st.button("🗑 초기화", key="notice_clear"):
        ss.notice_out = ""; st.rerun()

st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

# ── 영업일지 자동 생성 (달력 상담/매출 자동계산, 나머지는 직접 수정) ─
st.markdown("""
<div class="db-card">
  <div class="db-card-title">📈 영업일지 자동 생성</div>
  <div style='font-size:12px;color:var(--color-muted);margin-top:-10px;margin-bottom:14px'>
    상담건수·매출은 캘린더 기준으로 자동 채워집니다. 나머지 값은 아래 칸에서 바로 수정해서 쓰세요.
  </div>
</div>""", unsafe_allow_html=True)

if "rt_seq" not in st.session_state:
    st.session_state["rt_seq"] = 0
if st.button("🔄 캘린더 최신값으로 새로고침 (직접 수정한 내용은 초기화돼요)"):
    st.session_state["rt_seq"] += 1
    st.rerun()
rt_seq = st.session_state["rt_seq"]

today_consults = [c for c in consults if c["sched_date"] == today_str]
st.caption("상담 결과(등록/COD/미등록)는 위쪽 '🔵 상담 일정' 목록에서 체크하면 여기 자동 반영돼요.")

registered_ct = sum(1 for c in today_consults if c["result_status"] == "등록")
cod_ct = sum(1 for c in today_consults if c["result_status"] == "COD")
unregistered_ct = sum(1 for c in today_consults if c["result_status"] == "미등록")
auto_revenue = sum(c["actual_amount"] or c["expected_revenue"]
                    for c in today_consults if c["result_status"] in ("등록", "COD"))

rev_c1, rev_c2 = st.columns(2)
in_revenue = money_input(f"금일매출결과 (자동조회: {auto_revenue:,}원)",
                          log["actual_revenue"] or auto_revenue, "in_actual_revenue", rev_c1)
in_refund = money_input("환불", log["refund"], "in_refund", rev_c2)
if st.button("💾 금일매출결과·환불 저장"):
    run("UPDATE daily_log SET actual_revenue=?, refund=? WHERE log_date=?",
        (int(in_revenue), int(in_refund), today_str))
    st.rerun()

with st.expander("👤 담당자·통화시간 설정 (담당자는 고정, 통화시간은 매일 수정)"):
    n1, n2 = st.columns(2)
    in_rep1_name = n1.text_input("담당자1 이름", value=log["rep1_name"], key="in_rep1_name")
    in_rep2_name = n2.text_input("담당자2 이름", value=log["rep2_name"], key="in_rep2_name")
    t1c, t2c = st.columns(2)
    in_rep1_call = t1c.text_input("담당자1 통화시간", value=log["rep1_call"], placeholder="00:00:00", key="in_rep1_call")
    in_rep2_call = t2c.text_input("담당자2 통화시간", value=log["rep2_call"], placeholder="00:00:00", key="in_rep2_call")
    if st.button("💾 담당자·통화시간 저장"):
        run("""UPDATE daily_log SET rep1_name=?, rep2_name=?, rep1_call=?, rep2_call=? WHERE log_date=?""",
            (in_rep1_name.strip(), in_rep2_name.strip(), in_rep1_call.strip(), in_rep2_call.strip(), today_str))
        st.rerun()

cycle_start = get_state("cycle_start_date", today_str)
cycle_target_saved = int(get_state("cycle_target", "0") or 0)
cycle_base_saved = int(get_state("cycle_base_revenue", "0") or 0)

with st.expander("🆕 영업 시작 설정 (새 영업 사이클 시작할 때만)"):
    st.caption(f"현재 사이클 시작일: {cycle_start}")
    in_cycle_target = money_input("팀목표매출", cycle_target_saved, "in_cycle_target")
    in_cycle_base = money_input("현재달성매출 시작값", cycle_base_saved, "in_cycle_base")
    if st.button("🚀 이 값으로 새 영업 시작 (오늘부터 다시 쌓기 시작)"):
        set_state("cycle_start_date", today_str)
        set_state("cycle_target", str(int(in_cycle_target)))
        set_state("cycle_base_revenue", str(int(in_cycle_base)))
        st.rerun()

accumulated_since_start = q("SELECT COALESCE(SUM(actual_revenue),0) FROM daily_log WHERE log_date>=?",
                            (cycle_start,))[0][0]
month_target_live = cycle_target_saved
month_achieved_live = cycle_base_saved + accumulated_since_start
pct = round(month_achieved_live / month_target_live * 100) if month_target_live else 0

today_pct_lines = "\n".join(
    f"{c['expected_revenue'] // 10000}만원/{c['assignee'] or '미지정'}/{c['deposit_pct']}%"
    for c in sorted(today_consults, key=lambda c: c["sched_time"]))

morning_default = f"""{log['team_name']} 영업일지({today.month:02d}.{today.day:02d})

- 금일 입금예정: {today_rev // 10000}만원

{today_pct_lines}

- 금일 상담건수 : {today_cnt}건(정규{today_reg}건/단과{today_dan}건)
- 면접예정: {log['interview_count']}건"""

now_hm = now_kst.strftime("%H:%M")
pending_today = [c for c in today_consults if c["result_status"] not in ("등록", "COD", "미등록")]
ongoing_today = [c for c in pending_today if c["sched_time"] and c["sched_time"] <= now_hm]
upcoming_today = sorted((c for c in pending_today if c not in ongoing_today), key=lambda c: c["sched_time"])
ongoing_revenue = sum(c["expected_revenue"] for c in ongoing_today)
upcoming_lines = "\n".join(f"{c['sched_time']} {c['visit_type']} {c['expected_revenue'] // 10000}"
                           for c in upcoming_today)

pm3_values = {
    "팀명": log["team_name"],
    "입금완료": auto_revenue // 10000,
    "입금예정": today_rev // 10000,
    "상담중": ongoing_revenue // 10000,
    "상담목록": upcoming_lines,
    "익일상담": tmr_cnt, "익일예정": tmr_rev // 10000,
    "모레상담": daf_cnt, "모레예정": daf_rev // 10000,
    "익일면접": log["interview_count"],
    "따즈아분자": log["ddaz_num"], "따즈아분모": log["ddaz_den"],
}
pm3_tmpl = report_template_editor("15시보고", "tmpl_pm3", PM3_TEMPLATE_DEFAULT, pm3_values.keys())
pm3_default = render_report_template(pm3_tmpl, pm3_values)

close_values = {
    "팀명": log["team_name"],
    "상담건수": today_cnt, "등록": registered_ct, "COD": cod_ct, "미등록": unregistered_ct,
    "금일매출": f"{int(in_revenue):,}", "환불": f"{int(in_refund):,}",
    "담당자1": log["rep1_name"], "통화1": log["rep1_call"],
    "담당자2": log["rep2_name"], "통화2": log["rep2_call"],
    "익일상담": tmr_cnt, "익일매출": f"{tmr_rev:,}", "익일목표매출": f"{log['tmr_target']:,}",
    "월": datetime.strptime(cycle_start, "%Y-%m-%d").month,
    "팀목표매출": f"{month_target_live:,}", "현재달성매출": f"{month_achieved_live:,}", "달성율": pct,
}
close_tmpl = report_template_editor("마감보고", "tmpl_close", CLOSE_TEMPLATE_DEFAULT, close_values.keys())
close_default = render_report_template(close_tmpl, close_values)

# 보고서에 쓰인 원본 값이 바뀌면 아래 text_area가 자동으로 새로고침되도록,
# 위젯 key에 데이터 지문(fingerprint)을 포함시킴. (안 그러면 Streamlit이
# 이전에 렌더링된 위젯 값을 그대로 들고 있어서 값을 바꿔 저장해도
# 미리보기 텍스트가 안 바뀌는 문제가 있었음)
_fingerprint_src = "|".join(str(log.get(k, "")) for k in LOG_COLS)
_fingerprint_src += f"|{cycle_start}|{cycle_target_saved}|{cycle_base_saved}|{pm3_tmpl}|{close_tmpl}"
_fingerprint_src += "|" + "|".join(f"{c['id']}:{c['result_status']}:{c['actual_amount']}:{c['finalized']}"
                                    for c in consults)
rt_key = hashlib.md5(_fingerprint_src.encode()).hexdigest()[:10]

def _report_block(default_text, key_base, edit_height):
    """복사 우선(코드블록 우상단 복사 버튼) + 접이식 직접 수정."""
    edited = st.session_state.get(f"{key_base}_{rt_key}_{rt_seq}")
    st.code(edited if edited else default_text, language=None)
    with st.expander("✏️ 직접 수정"):
        st.text_area("직접 수정", value=default_text, height=edit_height,
                     key=f"{key_base}_{rt_key}_{rt_seq}", label_visibility="collapsed")

t1, t2, t3 = st.tabs(["출근보고", "15시보고", "마감보고"])
with t1:
    _report_block(morning_default, "rt_morning", 180)
with t2:
    _report_block(pm3_default, "rt_pm3", 140)
with t3:
    _report_block(close_default, "rt_close", 320)

st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

# ── 구글 시트 백업 ────────────────────────────────────────────
st.markdown("""
<div class="db-card">
  <div class="db-card-title">☁️ 구글 시트 백업</div>
  <div style='font-size:12px;color:var(--color-muted);margin-top:-10px;margin-bottom:14px'>
    하루 한 번 자동으로 상담·할일·일지 데이터를 백업합니다. DB는 그대로 sqlite를 씁니다.
  </div>
</div>""", unsafe_allow_html=True)

if not get_sheets_client():
    st.caption("아직 설정 안 됨 — Streamlit Secrets에 gcp_service_account / SHEETS_SPREADSHEET_ID를 추가하세요.")
else:
    last_backup = get_state("last_sheets_backup", "없음")
    st.caption(f"마지막 백업: {last_backup}")
    if st.button("지금 백업하기"):
        try:
            backup_to_sheets()
            set_state("last_sheets_backup", today_str)
            st.success("백업 완료!")
        except Exception as e:
            st.error(f"백업 실패: {e}")
